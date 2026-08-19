import json

import pytest

from openpyxl import Workbook

from scripts import db_ej_vers_xlsx
from scripts.construire_classeurs_751 import (
    Constructeur751,
    ajouter_feuille,
    feuille_copie_valeurs,
    lire_lignes_feuille,
    nouveau_classeur,
)
from shared.constantes import (
    ALIAS_COURT,
    NOM_COMPLET,
    NOMS_CLASSEURS_751,
    iterer_classeurs_751,
    resoudre_classeur_751,
)


def test_registre_developpe_18_classeurs_et_135_feuilles() -> None:
    classeurs = iterer_classeurs_751()

    assert len(classeurs) == 18
    assert sum(len(classeur.feuilles) for classeur in classeurs) == 135
    assert {classeur.nom_fichier for classeur in classeurs} == (
        db_ej_vers_xlsx.NOMS_CLASSEURS_ATTENDUS
    )


def test_registre_est_immuable() -> None:
    with pytest.raises(TypeError):
        NOMS_CLASSEURS_751["autre"] = NOMS_CLASSEURS_751["ej_entetes"]  # type: ignore[index]


def test_alias_sont_ordonnes_uniques_valides_et_completement_resolus() -> None:
    for classeur in iterer_classeurs_751():
        noms_complets = classeur.noms_feuilles(NOM_COMPLET)
        alias = classeur.noms_feuilles(ALIAS_COURT)

        assert len(noms_complets) == len(alias)
        assert len(alias) == len(set(alias))
        assert all(len(nom) <= 31 for nom in alias)
        assert all("{" not in nom and "}" not in nom for nom in (*noms_complets, *alias))


def test_registre_preserve_les_graphies_contractuelles_sensibles() -> None:
    entetes = resoudre_classeur_751("ej_entetes", boutique="MASSENA")
    lignes = resoudre_classeur_751("ej_lignes", boutique="MATURIN")
    z2_massena = resoudre_classeur_751("z2", boutique="MASSENA", annee=2024)
    z2_maturin = resoudre_classeur_751("z2", boutique="MATURIN", annee=2024)
    z1_massena = resoudre_classeur_751("z1", boutique="MASSENA", annee=2025)

    assert entetes.noms_feuilles(NOM_COMPLET)[:4] == (
        "ENTETES_TICKETS_MASSENA_0",
        "ENTETES_TICKETS_MASSENA_TriCrstNumInterne",
        "ENTETES_TICKETS_MASSENA_CtrlCoherenceEntete",
        "ENTETES_TICKETS_MASSENA_sequentialite",
    )
    assert "TD_OccurenceNumInterne" in entetes.noms_feuilles(NOM_COMPLET)
    assert "TD_OccurenceLibelleArticle" in lignes.noms_feuilles(NOM_COMPLET)
    assert z2_massena.noms_feuilles(NOM_COMPLET)[-2:] == (
        "Compare_Montant_Massena_Z2_ModeZZ1vsModeZZ2_2024",
        "Compare_Montant_MASSENA_Z2ModeZZ1vsEJ_2024",
    )
    assert z2_maturin.noms_feuilles(NOM_COMPLET)[-1] == (
        "Compare_Montant_MATURIN_Z2ModeZvsEJ_2024"
    )
    assert z1_massena.noms_feuilles(NOM_COMPLET)[-1] == (
        "Compare_Montant_MASSENA_Z1ModeZZ1vsEJ_2025"
    )
    assert resoudre_classeur_751("recettes_toutes").noms_feuilles(NOM_COMPLET) == (
        "recettes_mensuelles_tous_boutique_232425",
    )
    assert resoudre_classeur_751("comparaison_ca3").noms_feuilles(NOM_COMPLET) == (
        "CompareCA_Gesco_CA3",
    )


def test_variantes_conditionnelles_z2_sont_selectionnees_par_boutique() -> None:
    massena = resoudre_classeur_751("z2", boutique="MASSENA", annee=2023)
    maturin = resoudre_classeur_751("z2", boutique="MATURIN", annee=2023)

    assert len(massena.feuilles) == 8
    assert len(maturin.feuilles) == 7
    assert "Compare_ZZ1_vs_ZZ2" in massena.noms_feuilles(ALIAS_COURT)
    assert "Compare_ZZ1_vs_ZZ2" not in maturin.noms_feuilles(ALIAS_COURT)
    assert massena.noms_feuilles(ALIAS_COURT)[-1] == "Compare_ZZ1_vs_EJ"
    assert maturin.noms_feuilles(ALIAS_COURT)[-1] == "Compare_Z_vs_EJ"


def test_resolveur_refuse_un_contexte_invalide() -> None:
    with pytest.raises(ValueError, match="Boutique invalide"):
        resoudre_classeur_751("ej_entetes", boutique="INCONNUE")
    with pytest.raises(ValueError, match="Année invalide"):
        resoudre_classeur_751("z1", boutique="MASSENA", annee=2022)
    with pytest.raises(ValueError, match="n'accepte pas de boutique"):
        resoudre_classeur_751("recettes_toutes", boutique="MASSENA")


def test_constructeur_xlsx_refuse_noms_longs_et_doublons() -> None:
    classeur = Workbook()
    with pytest.raises(ValueError, match="trop long"):
        ajouter_feuille(classeur, "X" * 32, ["A"], [], "TropLong")

    ajouter_feuille(classeur, "Court", ["A"], [], "Court")
    with pytest.raises(ValueError, match="dupliqué"):
        ajouter_feuille(classeur, "Court", ["A"], [], "Doublon")


def test_manifeste_expose_la_correspondance_bijective(tmp_path) -> None:
    definition = resoudre_classeur_751("comparaison_ca3")
    constructeur = object.__new__(Constructeur751)
    constructeur.xlsx = tmp_path
    constructeur.manifeste = {"workbooks": [], "checks": []}
    classeur = nouveau_classeur()
    ajouter_feuille(
        classeur,
        definition.noms_feuilles(ALIAS_COURT)[0],
        ["A"],
        [{"A": 1}],
        "TestCorrespondance",
    )

    constructeur.exporter(classeur, definition)

    entree = constructeur.manifeste["workbooks"][0]
    assert entree["sheetNameMode"] == ALIAS_COURT
    assert entree["sheets"] == ["CompareCA_Gesco_CA3"]
    assert entree["requestedSheets"] == ["CompareCA_Gesco_CA3"]
    assert entree["sheetNameMappings"] == [{
        "nomComplet": "CompareCA_Gesco_CA3",
        "aliasCourt": "CompareCA_Gesco_CA3",
        "nomProduit": "CompareCA_Gesco_CA3",
    }]
    json.dumps(constructeur.manifeste, ensure_ascii=False)


def test_copie_en_valeur_lit_la_feuille_source_sans_formule_de_liaison() -> None:
    classeur = nouveau_classeur()
    source = ajouter_feuille(
        classeur,
        "Source",
        ["E_NUM_INTERNE", "E_TTC"],
        [{"E_NUM_INTERNE": "000002", "E_TTC": 120}],
        "SourceValeurs",
    )

    cible = feuille_copie_valeurs(
        classeur,
        "Cible",
        source,
        ["E_NUM_INTERNE", "E_TTC"],
        "CibleValeurs",
    )

    assert lire_lignes_feuille(cible) == [{"E_NUM_INTERNE": "000002", "E_TTC": 120}]
    assert all(
        not (isinstance(cellule.value, str) and cellule.value.startswith("="))
        for ligne in cible.iter_rows(min_row=2)
        for cellule in ligne
    )


def test_classeur_entetes_respecte_copies_et_formules_contractuelles() -> None:
    constructeur = object.__new__(Constructeur751)
    constructeur.donnees = {"entetes": {"MASSENA": [{
        "nomfichier": "EJ010123.txt", "E_NUM_INTERNE": "000001", "E_NUM_TICKET": "000001",
        "E_DATE_TICKET": "2023-01-01", "E_HEURE_TICKET": "10:00:00",
        "E_HT1": 10, "E_HT2": None, "E_HT3": -1, "E_HT4": 0,
        "E_TVA1": -2, "E_TVA2": 0, "E_TVA3": 0, "E_TVA4": 0,
        "E_HT_NON_TAXABLE": 0, "E_TTC": 12,
        "E_MDP_CB": 12, "E_MDP_ESPECES": 0, "E_MDP_CHEQUES": 0,
    }]}}
    classeur_produit = []
    constructeur.exporter = lambda classeur, _definition: classeur_produit.append(classeur)

    constructeur.construire_entetes("MASSENA")

    entetes = [cellule.value for cellule in classeur_produit[0]["Entetes_sequentialite"][1]]
    assert entetes == [
        "nomfichier", "E_NUM_INTERNE", "E_NUM_TICKET", "E_DATE_TICKET",
        "E_HEURE_TICKET", "AJ_TROU_NUM_TICKET",
    ]
    assert lire_lignes_feuille(classeur_produit[0]["DoublonNumInterne"]) == (
        lire_lignes_feuille(classeur_produit[0]["TD_OccurenceNumInterne"])
    )
    assert lire_lignes_feuille(classeur_produit[0]["DoublonNumTicket"]) == (
        lire_lignes_feuille(classeur_produit[0]["TD_OccurenceNumTicket"])
    )

    complete = classeur_produit[0]["Entetes_CplteAnneeMois"]
    entetes_complete = [cellule.value for cellule in complete[1]]
    colonne_total_ht = entetes_complete.index("AJ_TOTAL_HT") + 1
    colonne_total_tva = entetes_complete.index("AJ_TOTAL_TVA_20") + 1
    assert complete.cell(2, colonne_total_ht).value == "=F2+G2+H2+I2+N2"
    assert complete.cell(2, colonne_total_tva).value == "=J2"

    recette_janvier = lire_lignes_feuille(
        classeur_produit[0]["TD_TotalHtTvaTtc"]
    )[0]
    assert recette_janvier["SOMME_AJ_TOTAL_HT"] == 9
    assert recette_janvier["SOMME_AJ_TOTAL_TVA_20"] == -2
    assert constructeur.filiations_actives[5]["operation"] == "copie en valeur"
    assert constructeur.filiations_actives[7]["operation"] == "copie en valeur"
    assert constructeur.filiations_actives[8] == {
        "targetSheet": "Entetes_CplteAnneeMois",
        "immediateSources": ["Entetes_TriNumInterne"],
        "operation": "copie, enrichissement année/mois et formules HT/TVA",
    }


def test_coherence_entete_ligne_utilise_une_formule_excel() -> None:
    constructeur = object.__new__(Constructeur751)
    constructeur.donnees = {"lignes": {"MASSENA": [{
        "nomfichier": "EJ010123.txt",
        "E_NUM_INTERNE": "000001",
        "E_NUM_TICKET": "000001",
        "E_DATE_TICKET": "2023-01-01",
        "E_HEURE_TICKET": "10:00:00",
        "E_TTC": 10,
        "D_QUANTITE_ARTICLE": 1,
        "D_LIBELLE_ARTICLE": "",
        "D_TAUX_TVA_ARTICLE": 20,
        "D_MONTANT_ARTICLE": 12,
        "D_CORRECTION": -2,
        "D_AUTRE_INFO": "",
    }]}}
    classeur_produit = []
    constructeur.exporter = lambda classeur, _definition: classeur_produit.append(classeur)

    constructeur.construire_lignes("MASSENA")

    coherence = classeur_produit[0]["CtrlCoherence_EnteteLigne"]
    assert coherence["F2"].value == "=B2-(D2+E2)"
    assert constructeur.filiations_actives[4] == {
        "targetSheet": "CtrlCoherence_EnteteLigne",
        "immediateSources": ["TD_TotalLignesParTicket"],
        "operation": "copie en valeur et formule d'écart",
    }


def test_manifeste_enregistre_la_source_immediate(tmp_path) -> None:
    definition = resoudre_classeur_751("comparaison_ca3")
    constructeur = object.__new__(Constructeur751)
    constructeur.xlsx = tmp_path
    constructeur.manifeste = {"workbooks": [], "checks": []}
    constructeur.filiations_actives = [{
        "targetSheet": "CompareCA_Gesco_CA3",
        "immediateSources": ["recettes_mensuelles_tous_boutique_232425.xlsx#recettes_mensuelles_tous"],
        "operation": "copie des recettes reconstituées et comparaison CA3",
    }]
    classeur = nouveau_classeur()
    ajouter_feuille(classeur, "CompareCA_Gesco_CA3", ["A"], [{"A": 1}], "TestFiliation")

    constructeur.exporter(classeur, definition)

    assert constructeur.manifeste["workbooks"][0]["sheetLineage"] == constructeur.filiations_actives
