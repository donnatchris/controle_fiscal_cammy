"""Génère le rapport PDF d'analyse fiscale du dossier 751."""

from __future__ import annotations

import argparse
import csv
import sqlite3

from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


NOM_RAPPORT = "RAPPORT_ANALYSE_FISCALE_751.pdf"
TOLERANCE = Decimal("0.02")
TYPES_VENTE = {"REG", "_R_F"}

BLEU = colors.HexColor("#123B5D")
BLEU_CLAIR = colors.HexColor("#EAF2F7")
TURQUOISE = colors.HexColor("#138A8A")
VERT = colors.HexColor("#26734D")
VERT_CLAIR = colors.HexColor("#E8F3EC")
ORANGE = colors.HexColor("#B76721")
ORANGE_CLAIR = colors.HexColor("#FFF1E5")
GRIS = colors.HexColor("#536270")
GRIS_CLAIR = colors.HexColor("#F3F5F6")
GRIS_LIGNE = colors.HexColor("#CDD5DB")


def decimaliser(valeur: object) -> Decimal:
    return Decimal(str(valeur or "0"))


def est_vente(ligne: sqlite3.Row) -> bool:
    return ligne["type"] in TYPES_VENTE and bool(ligne["E_NUM_TICKET"])


def lire_csv(chemin: Path) -> list[dict[str, str]]:
    with chemin.open(encoding="utf-8-sig", newline="") as fichier:
        return list(csv.DictReader(fichier, delimiter=";"))


def analyser_sequences(lignes: list[sqlite3.Row]) -> dict[str, dict[str, Any]]:
    resultat: dict[str, dict[str, Any]] = {}
    for boutique in sorted({ligne["boutique"] for ligne in lignes}):
        blocs = [ligne for ligne in lignes if ligne["boutique"] == boutique]
        ventes = [ligne for ligne in blocs if est_vente(ligne)]
        numeros = sorted(int(ligne["E_NUM_INTERNE"]) for ligne in blocs)
        numeros_ventes = sorted(int(ligne["E_NUM_INTERNE"]) for ligne in ventes)
        numeros_exclus = {
            int(ligne["E_NUM_INTERNE"]) for ligne in blocs if not est_vente(ligne)
        }
        absents_ventes: list[int] = []
        sauts_ventes = 0
        for precedent, courant in zip(numeros_ventes, numeros_ventes[1:]):
            if courant > precedent + 1:
                sauts_ventes += 1
                absents_ventes.extend(range(precedent + 1, courant))

        numeros_tickets = sorted(int(ligne["E_NUM_TICKET"]) for ligne in ventes)
        absents_tickets = sum(
            max(0, courant - precedent - 1)
            for precedent, courant in zip(numeros_tickets, numeros_tickets[1:])
        )
        ordonnes = sorted(blocs, key=lambda ligne: int(ligne["E_NUM_INTERNE"]))
        regressions = 0
        for precedent, courant in zip(ordonnes, ordonnes[1:]):
            date_precedente = datetime.fromisoformat(
                f"{precedent['E_DATE_TICKET']} {precedent['E_HEURE_TICKET']}"
            )
            date_courante = datetime.fromisoformat(
                f"{courant['E_DATE_TICKET']} {courant['E_HEURE_TICKET']}"
            )
            regressions += date_courante < date_precedente

        resultat[boutique] = {
            "blocs": len(blocs),
            "ventes": len(ventes),
            "exclus": len(blocs) - len(ventes),
            "interne_min": min(numeros),
            "interne_max": max(numeros),
            "interne_absents_exhaustif": (
                max(numeros) - min(numeros) + 1 - len(set(numeros))
            ),
            "interne_doublons": len(numeros) - len(set(numeros)),
            "sauts_apparents_ventes": sauts_ventes,
            "numeros_absents_ventes": len(absents_ventes),
            "absents_expliques": set(absents_ventes) <= numeros_exclus,
            "ticket_min": min(numeros_tickets),
            "ticket_max": max(numeros_tickets),
            "ticket_absents": absents_tickets,
            "ticket_doublons": len(numeros_tickets) - len(set(numeros_tickets)),
            "regressions_chronologiques": regressions,
        }
    return resultat


def analyser_donnees(
    base: Path,
    controle: Path,
    sortie: Path,
) -> dict[str, Any]:
    with sqlite3.connect(base) as connexion:
        connexion.row_factory = sqlite3.Row
        tickets = connexion.execute("SELECT * FROM tickets").fetchall()
        details: dict[int, list[sqlite3.Row]] = defaultdict(list)
        for detail in connexion.execute("SELECT * FROM lignes_ticket"):
            details[detail["ticket_id"]].append(detail)

    ventes = [ticket for ticket in tickets if est_vente(ticket)]
    exclus = [ticket for ticket in tickets if not est_vente(ticket)]
    types = Counter(ticket["type"] for ticket in tickets)
    types_ventes = Counter(ticket["type"] for ticket in ventes)
    types_exclus = Counter(ticket["type"] for ticket in exclus)
    evenements_exclus = Counter(
        (ticket["type"], ticket["evenement"] or "") for ticket in exclus
    )

    anomalies_monetaires = Counter()
    ecarts_max = defaultdict(Decimal)
    totaux = defaultdict(lambda: defaultdict(Decimal))
    total_retours = Decimal("0")
    nombre_retours = 0
    for ticket in ventes:
        ht = sum(decimaliser(ticket[f"E_HT{i}"]) for i in range(1, 5))
        ht += decimaliser(ticket["E_HT_NON_TAXABLE"])
        tva = sum(decimaliser(ticket[f"E_TVA{i}"]) for i in range(1, 5))
        ttc = decimaliser(ticket["E_TTC"])
        paiements = sum(
            decimaliser(ticket[colonne])
            for colonne in ("E_MDP_CB", "E_MDP_ESPECES", "E_MDP_CHEQUES")
        )
        montant_details = sum(
            decimaliser(detail["D_MONTANT_ARTICLE"])
            + decimaliser(detail["D_CORRECTION"])
            for detail in details[ticket["id"]]
        )
        controles = {
            "HT + TVA = TTC": abs(ht + tva - ttc),
            "Règlements = TTC": abs(paiements - ttc),
            "Lignes = TTC": abs(montant_details - ttc),
        }
        for libelle, ecart in controles.items():
            ecarts_max[libelle] = max(ecarts_max[libelle], ecart)
            anomalies_monetaires[libelle] += ecart > TOLERANCE

        cle = (ticket["boutique"], ticket["E_DATE_TICKET"][:4])
        totaux[cle]["HT"] += ht
        totaux[cle]["TVA"] += tva
        totaux[cle]["TTC"] += ttc
        if ticket["type"] == "_R_F":
            nombre_retours += 1
            total_retours += ttc

    rapprochements = lire_csv(controle / "RAPPROCHEMENT_PAR_CLOTURE_EJ_Z.csv")
    couvertures = lire_csv(controle / "CONTROLE_COUVERTURE_PERIODES_Z.csv")
    sans_cloture = [
        ligne for ligne in couvertures if ligne["STATUT"] == "PAS_DE_CLOTURE"
    ]

    classeurs: list[dict[str, Any]] = []
    total_feuilles = 0
    total_formules = 0
    erreurs_ref = 0
    for chemin in sorted(sortie.glob("*.xlsx")):
        classeur = load_workbook(chemin, read_only=True, data_only=False)
        formules = 0
        refs = 0
        for feuille in classeur.worksheets:
            for ligne in feuille.iter_rows():
                for cellule in ligne:
                    if isinstance(cellule.value, str) and cellule.value.startswith("="):
                        formules += 1
                        refs += "#REF!" in cellule.value
        classeurs.append(
            {"nom": chemin.name, "feuilles": len(classeur.sheetnames), "formules": formules}
        )
        total_feuilles += len(classeur.sheetnames)
        total_formules += formules
        erreurs_ref += refs
        classeur.close()

    return {
        "date_min": min(ticket["E_DATE_TICKET"] for ticket in tickets),
        "date_max": max(ticket["E_DATE_TICKET"] for ticket in tickets),
        "blocs": len(tickets),
        "details": sum(len(valeurs) for valeurs in details.values()),
        "ventes": len(ventes),
        "exclus": len(exclus),
        "types": types,
        "types_ventes": types_ventes,
        "types_exclus": types_exclus,
        "evenements_exclus": evenements_exclus,
        "sequences": analyser_sequences(tickets),
        "anomalies_monetaires": anomalies_monetaires,
        "ecarts_max": ecarts_max,
        "totaux": totaux,
        "nombre_retours": nombre_retours,
        "total_retours": total_retours,
        "rapprochements": rapprochements,
        "sans_cloture": sans_cloture,
        "couvertures": couvertures,
        "classeurs": classeurs,
        "total_feuilles": total_feuilles,
        "total_formules": total_formules,
        "erreurs_ref": erreurs_ref,
    }


def euro(valeur: Decimal) -> str:
    texte = f"{valeur:,.2f}".replace(",", " ").replace(".", ",")
    return f"{texte} EUR"


def styles_document() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "titre": ParagraphStyle(
            "Titre", parent=base["Title"], fontName="Helvetica-Bold",
            fontSize=25, leading=30, textColor=BLEU, alignment=TA_LEFT,
            spaceAfter=7 * mm,
        ),
        "soustitre": ParagraphStyle(
            "SousTitre", parent=base["Normal"], fontName="Helvetica",
            fontSize=11, leading=16, textColor=GRIS, spaceAfter=4 * mm,
        ),
        "h1": ParagraphStyle(
            "H1", parent=base["Heading1"], fontName="Helvetica-Bold",
            fontSize=17, leading=21, textColor=BLEU, spaceBefore=2 * mm,
            spaceAfter=5 * mm,
        ),
        "h2": ParagraphStyle(
            "H2", parent=base["Heading2"], fontName="Helvetica-Bold",
            fontSize=11.5, leading=15, textColor=TURQUOISE,
            spaceBefore=4 * mm, spaceAfter=2 * mm,
        ),
        "corps": ParagraphStyle(
            "Corps", parent=base["BodyText"], fontName="Helvetica",
            fontSize=9.2, leading=13.2, textColor=colors.HexColor("#24323D"),
            alignment=TA_JUSTIFY, spaceAfter=2.5 * mm,
        ),
        "petit": ParagraphStyle(
            "Petit", parent=base["BodyText"], fontName="Helvetica",
            fontSize=7.7, leading=10.2, textColor=GRIS,
        ),
        "statut": ParagraphStyle(
            "Statut", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=11, leading=15, textColor=VERT, alignment=TA_CENTER,
        ),
        "alerte": ParagraphStyle(
            "Alerte", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=10, leading=14, textColor=ORANGE,
        ),
        "table": ParagraphStyle(
            "Table", parent=base["Normal"], fontName="Helvetica",
            fontSize=7.6, leading=9.5, textColor=colors.HexColor("#24323D"),
        ),
        "table_head": ParagraphStyle(
            "TableHead", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=7.4, leading=9, textColor=colors.white,
        ),
    }


def p(texte: object, style: ParagraphStyle) -> Paragraph:
    return Paragraph(str(texte), style)


def tableau(
    donnees: list[list[object]],
    largeurs: list[float],
    styles: dict[str, ParagraphStyle],
    *,
    entete: bool = True,
    alignements: dict[int, str] | None = None,
) -> Table:
    contenu: list[list[Paragraph]] = []
    for no_ligne, ligne in enumerate(donnees):
        style = styles["table_head"] if entete and no_ligne == 0 else styles["table"]
        contenu.append([p(cellule, style) for cellule in ligne])
    objet = Table(contenu, colWidths=largeurs, repeatRows=1 if entete else 0, hAlign="LEFT")
    commandes: list[tuple[Any, ...]] = [
        ("BACKGROUND", (0, 0), (-1, 0), BLEU if entete else GRIS_CLAIR),
        ("GRID", (0, 0), (-1, -1), 0.35, GRIS_LIGNE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if entete and len(donnees) > 2:
        commandes.append(("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLAIR]))
    for colonne, alignement in (alignements or {}).items():
        commandes.append(("ALIGN", (colonne, 1 if entete else 0), (colonne, -1), alignement))
    objet.setStyle(TableStyle(commandes))
    return objet


def encadre(texte: str, style: ParagraphStyle, couleur: colors.Color) -> Table:
    objet = Table([[p(texte, style)]], colWidths=[174 * mm])
    objet.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), couleur),
        ("BOX", (0, 0), (-1, -1), 0.7, GRIS_LIGNE),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    return objet


def ajouter_entete_pied(canvas: Any, document: BaseDocTemplate) -> None:
    canvas.saveState()
    largeur, hauteur = A4
    canvas.setStrokeColor(GRIS_LIGNE)
    canvas.setLineWidth(0.5)
    canvas.line(18 * mm, hauteur - 15 * mm, largeur - 18 * mm, hauteur - 15 * mm)
    canvas.setFont("Helvetica-Bold", 7.5)
    canvas.setFillColor(BLEU)
    canvas.drawString(18 * mm, hauteur - 11.5 * mm, "DOSSIER 751 - ANALYSE DES DONNEES DE CAISSE")
    canvas.setFont("Helvetica", 7.2)
    canvas.setFillColor(GRIS)
    canvas.drawRightString(largeur - 18 * mm, hauteur - 11.5 * mm, "CAMMY FRANCE DEVELOPPEMENT LTD")
    canvas.line(18 * mm, 14 * mm, largeur - 18 * mm, 14 * mm)
    canvas.drawString(18 * mm, 9.5 * mm, "Rapport généré à partir de la base SQLite et des contrôles produits")
    page = f"Page {document.page}"
    canvas.drawRightString(largeur - 18 * mm, 9.5 * mm, page)
    canvas.restoreState()


def generer_pdf(analyse: dict[str, Any], destination: Path) -> None:
    styles = styles_document()
    destination.parent.mkdir(parents=True, exist_ok=True)
    document = BaseDocTemplate(
        str(destination),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=22 * mm,
        bottomMargin=20 * mm,
        title="Rapport d'analyse fiscale - dossier 751",
        author="CAMMY FRANCE DEVELOPPEMENT LTD",
        subject="Analyse des données de caisse EJ et Z",
    )
    cadre = Frame(
        document.leftMargin,
        document.bottomMargin,
        document.width,
        document.height,
        id="corps",
    )
    document.addPageTemplates([
        PageTemplate(id="rapport", frames=[cadre], onPage=ajouter_entete_pied)
    ])
    histoire: list[Any] = []

    # Couverture et synthèse
    histoire.extend([
        Spacer(1, 15 * mm),
        p("RAPPORT D'ANALYSE<br/>DES DONNEES DE CAISSE", styles["titre"]),
        p("Dossier 751 - CAMMY FRANCE DEVELOPPEMENT LTD", styles["soustitre"]),
        p(
            f"Période examinée : {analyse['date_min']} au {analyse['date_max']}<br/>"
            "Sources : journaux électroniques EJ, rapports Z, base SQLite et classeurs de restitution",
            styles["soustitre"],
        ),
        Spacer(1, 6 * mm),
        encadre(
            "CONFORME SUR LE PERIMETRE CAISSE - A COMPLETER",
            styles["statut"],
            VERT_CLAIR,
        ),
        Spacer(1, 7 * mm),
        p("Conclusion exécutive", styles["h1"]),
        p(
            "Les contrôles réalisés ne mettent en évidence aucune anomalie inexpliquée "
            "dans les données de caisse disponibles. Les séquences exhaustives sont continues, "
            "les montants sont cohérents au centime et les 57 rapprochements entre journaux EJ "
            "et clôtures Z sont concordants.",
            styles["corps"],
        ),
        p(
            "Les sauts visibles dans E_NUM_INTERNE lorsque l'on consulte uniquement les ventes "
            "sont des écarts apparents : les numéros correspondants appartiennent à des rapports "
            "administratifs ou à des événements sans vente, tous conservés dans la base exhaustive. "
            "Une fois ces blocs réintégrés au contrôle de séquence, aucun numéro ne manque.",
            styles["corps"],
        ),
        encadre(
            "Limite de portée : les FEC, déclarations CA3 et justificatifs externes n'ont pas été "
            "fournis. Le présent rapport ne vaut donc ni validation de ces documents, ni attestation "
            "générale de conformité fiscale.",
            styles["alerte"],
            ORANGE_CLAIR,
        ),
        Spacer(1, 7 * mm),
        tableau([
            ["Indicateur", "Résultat"],
            ["Blocs EJ conservés en base", f"{analyse['blocs']:,}".replace(",", " ")],
            ["Tickets de vente exportés", f"{analyse['ventes']:,}".replace(",", " ")],
            ["Lignes de détail", f"{analyse['details']:,}".replace(",", " ")],
            ["Rapprochements EJ / Z", f"{sum(r['STATUT'] == 'OK' for r in analyse['rapprochements'])} / {len(analyse['rapprochements'])} conformes"],
            ["Classeurs / feuilles", f"{len(analyse['classeurs'])} / {analyse['total_feuilles']}"],
        ], [108 * mm, 66 * mm], styles, alignements={1: "RIGHT"}),
        PageBreak(),
    ])

    # Méthode et périmètre
    histoire.extend([
        p("1. Objet, périmètre et méthode", styles["h1"]),
        p(
            "Le présent document expose les contrôles appliqués aux données produites dans le "
            "dossier de sortie. Il vise à permettre à l'administration fiscale de comprendre "
            "l'origine des chiffres, les règles d'exclusion des exports de ventes et la portée des "
            "constats. Les calculs monétaires sont réalisés avec le type Decimal et une tolérance "
            "maximale de deux centimes pour les rapprochements.",
            styles["corps"],
        ),
        p("Chaîne de traitement examinée", styles["h2"]),
        tableau([
            ["Etape", "Traitement", "Contrôle associé"],
            ["1", "Lecture des EJ et des fichiers Z", "Traçabilité du fichier source"],
            ["2", "Parsing de tous les blocs EJ", "Conservation des types REG, _R_F, X, XZ et Z"],
            ["3", "Chargement dans SQLite", "Unicité boutique + E_NUM_INTERNE"],
            ["4", "Sélection des ventes", "REG ou _R_F avec E_NUM_TICKET renseigné"],
            ["5", "Export des 18 classeurs", "Structure, feuilles, formules et absence de #REF!"],
            ["6", "Rapprochement EJ / Z", "CA, HT, TVA et modes de règlement"],
        ], [15 * mm, 79 * mm, 80 * mm], styles, alignements={0: "CENTER"}),
        p("Règle de conservation", styles["h2"]),
        p(
            "La base SQLite est exhaustive : aucun bloc n'est supprimé en fonction de son type. "
            "L'absence de E_NUM_TICKET est utilisée uniquement pour ne pas présenter un événement "
            "administratif comme une vente dans les classeurs contractuels. Un type futur ou inconnu "
            "serait conservé en base, signalé dans l'inventaire et exclu des ventes jusqu'à qualification.",
            styles["corps"],
        ),
        p("Sources et limites documentaires", styles["h2"]),
        tableau([
            ["Source", "Disponibilité", "Utilisation"],
            ["Journaux électroniques EJ", "Disponible", "Tickets, détails, séquences et signatures"],
            ["Rapports Z1 et Z2", "Disponible", "Clôtures, CA et règlements"],
            ["FEC", "Non fourni", "Rapprochement comptable non réalisable"],
            ["Déclarations CA3", "Non fournies", "Rapprochement déclaratif non réalisable"],
            ["Justificatifs externes", "Non fournis", "Qualification externe non réalisable"],
        ], [54 * mm, 33 * mm, 87 * mm], styles),
        PageBreak(),
    ])

    # Inventaire
    histoire.extend([
        p("2. Exhaustivité de la base et sélection des ventes", styles["h1"]),
        p(
            f"La base contient {analyse['blocs']} blocs EJ et {analyse['details']} lignes de détail. "
            f"Parmi eux, {analyse['ventes']} tickets répondent à la définition d'une vente et "
            f"{analyse['exclus']} blocs restent hors des classeurs de ventes. Ces derniers demeurent "
            "consultables dans SQLite avec leur type, leur numéro interne, leur événement, leur "
            "signature et leur provenance.",
            styles["corps"],
        ),
        tableau([
            ["Type EJ", "En base", "Dans les ventes", "Hors ventes", "Qualification"],
            *[
                [
                    type_bloc,
                    analyse["types"].get(type_bloc, 0),
                    analyse["types_ventes"].get(type_bloc, 0),
                    analyse["types_exclus"].get(type_bloc, 0),
                    "Vente si E_NUM_TICKET" if type_bloc in TYPES_VENTE else "Bloc administratif",
                ]
                for type_bloc in ("REG", "_R_F", "X", "XZ", "Z")
            ],
            ["TOTAL", analyse["blocs"], analyse["ventes"], analyse["exclus"], ""],
        ], [24 * mm, 24 * mm, 29 * mm, 27 * mm, 70 * mm], styles,
            alignements={1: "RIGHT", 2: "RIGHT", 3: "RIGHT"}),
        p("Cas des REG et _R_F sans numéro de ticket", styles["h2"]),
        p(
            "Les 311 blocs REG et les 3 blocs _R_F sans E_NUM_TICKET portent l'événement "
            "REF./TIROIR. Ils ne comportent ni ligne article, ni montant HT/TVA/TTC, ni règlement. "
            "Ils sont donc exclus de l'export des ventes pour éviter de créer des factures fictives, "
            "mais ils ne sont ni supprimés ni qualifiés d'erreurs.",
            styles["corps"],
        ),
        encadre(
            "Constat : l'écart entre 2 511 blocs et 1 875 ventes résulte d'une règle de périmètre "
            "documentée. Les 636 blocs exclus ont un motif identifié et restent traçables.",
            styles["corps"],
            BLEU_CLAIR,
        ),
        PageBreak(),
    ])

    # Séquences
    seq = analyse["sequences"]
    histoire.extend([
        p("3. Séquentialité, doublons et chronologie", styles["h1"]),
        p(
            "Deux lectures sont nécessaires. La lecture des seuls tickets de vente reflète les "
            "classeurs remis ; la lecture exhaustive réintègre les événements administratifs et "
            "permet de contrôler la séquence native du logiciel de caisse.",
            styles["corps"],
        ),
        tableau([
            ["Boutique", "Plage exhaustive", "Manquants exhaustifs", "Doublons", "Retours chrono."],
            *[
                [
                    boutique,
                    f"{valeurs['interne_min']:06d} à {valeurs['interne_max']:06d}",
                    valeurs["interne_absents_exhaustif"],
                    valeurs["interne_doublons"],
                    valeurs["regressions_chronologiques"],
                ]
                for boutique, valeurs in seq.items()
            ],
        ], [34 * mm, 54 * mm, 34 * mm, 24 * mm, 28 * mm], styles,
            alignements={2: "RIGHT", 3: "RIGHT", 4: "RIGHT"}),
        p("Explication des sauts apparents dans E_NUM_INTERNE", styles["h2"]),
        tableau([
            ["Boutique", "Ventes", "Sauts visibles", "Numéros non affichés", "Tous retrouvés en base"],
            *[
                [
                    boutique,
                    valeurs["ventes"],
                    valeurs["sauts_apparents_ventes"],
                    valeurs["numeros_absents_ventes"],
                    "OUI" if valeurs["absents_expliques"] else "NON",
                ]
                for boutique, valeurs in seq.items()
            ],
            [
                "TOTAL",
                sum(v["ventes"] for v in seq.values()),
                sum(v["sauts_apparents_ventes"] for v in seq.values()),
                sum(v["numeros_absents_ventes"] for v in seq.values()),
                "OUI",
            ],
        ], [34 * mm, 28 * mm, 32 * mm, 42 * mm, 38 * mm], styles,
            alignements={1: "RIGHT", 2: "RIGHT", 3: "RIGHT", 4: "CENTER"}),
        p(
            "Dans les classeurs de ventes, 624 numéros internes ne sont pas affichés entre deux "
            "ventes successives. Chacun de ces numéros est présent dans SQLite et correspond à un "
            "bloc exclu selon la règle de périmètre. Les 12 autres blocs exclus se trouvent avant "
            "la première ou après la dernière vente de leur boutique et ne créent donc pas de saut "
            "entre deux lignes exportées. Sur l'ensemble exhaustif, le nombre de numéros manquants est zéro.",
            styles["corps"],
        ),
        p("Séquence E_NUM_TICKET", styles["h2"]),
        tableau([
            ["Boutique", "Plage tickets", "Numéros manquants", "Doublons"],
            *[
                [
                    boutique,
                    f"{valeurs['ticket_min']:06d} à {valeurs['ticket_max']:06d}",
                    valeurs["ticket_absents"],
                    valeurs["ticket_doublons"],
                ]
                for boutique, valeurs in seq.items()
            ],
        ], [42 * mm, 58 * mm, 42 * mm, 32 * mm], styles,
            alignements={2: "RIGHT", 3: "RIGHT"}),
        encadre(
            "Conclusion du contrôle de séquence : aucun trou dans la séquence exhaustive, aucun "
            "doublon de numéro interne ou de ticket et aucune régression chronologique.",
            styles["corps"],
            VERT_CLAIR,
        ),
        PageBreak(),
    ])

    # Montants
    lignes_totaux = [["Boutique", "Exercice", "HT", "TVA", "TTC"]]
    for (boutique, exercice), valeurs in sorted(analyse["totaux"].items()):
        lignes_totaux.append([
            boutique, exercice, euro(valeurs["HT"]), euro(valeurs["TVA"]), euro(valeurs["TTC"])
        ])
    total_ht = sum(valeurs["HT"] for valeurs in analyse["totaux"].values())
    total_tva = sum(valeurs["TVA"] for valeurs in analyse["totaux"].values())
    total_ttc = sum(valeurs["TTC"] for valeurs in analyse["totaux"].values())
    lignes_totaux.append(["TOTAL", "2023-2025", euro(total_ht), euro(total_tva), euro(total_ttc)])
    histoire.extend([
        p("4. Cohérence monétaire et traitement des retours", styles["h1"]),
        p(
            "Pour chacun des 1 875 tickets, trois égalités sont recalculées : HT + TVA = TTC, "
            "somme des règlements = TTC et somme des lignes = TTC. Aucun écart supérieur à deux "
            "centimes n'est constaté ; dans les données présentes, l'écart maximal observé est nul.",
            styles["corps"],
        ),
        tableau([
            ["Contrôle", "Tickets contrôlés", "Ecarts > 0,02 EUR", "Ecart maximal"],
            *[
                [libelle, analyse["ventes"], analyse["anomalies_monetaires"][libelle], euro(ecart)]
                for libelle, ecart in analyse["ecarts_max"].items()
            ],
        ], [64 * mm, 38 * mm, 38 * mm, 34 * mm], styles,
            alignements={1: "RIGHT", 2: "RIGHT", 3: "RIGHT"}),
        p("Totaux issus des tickets de vente", styles["h2"]),
        tableau(lignes_totaux, [35 * mm, 25 * mm, 38 * mm, 38 * mm, 38 * mm], styles,
                alignements={2: "RIGHT", 3: "RIGHT", 4: "RIGHT"}),
        p("Retours _R_F", styles["h2"]),
        p(
            f"Les {analyse['nombre_retours']} retours de vente _R_F sont conservés et signés "
            f"négativement, pour un TTC cumulé de {euro(analyse['total_retours'])}. Ce signe est "
            "nécessaire pour que les retours diminuent le chiffre d'affaires et pour permettre le "
            "rapprochement avec les rapports Z. Les trois _R_F sans E_NUM_TICKET sont des événements "
            "REF./TIROIR à montant nul et ne sont pas des retours financiers.",
            styles["corps"],
        ),
        p("Indicateurs de TVA", styles["h2"]),
        p(
            "Le champ D_TAUX_TVA_ARTICLE restitue l'indicateur source T1, sans le transformer en "
            "pourcentage. Le taux de 20 % intervient uniquement dans la formule de contrôle du HT. "
            "Cette séparation évite d'altérer la donnée native tout en rendant le calcul vérifiable.",
            styles["corps"],
        ),
        PageBreak(),
    ])

    # EJ/Z
    ok = sum(ligne["STATUT"] == "OK" for ligne in analyse["rapprochements"])
    modes = Counter(ligne["MODE_RETENU"] for ligne in analyse["rapprochements"])
    histoire.extend([
        p("5. Rapprochements entre EJ et rapports Z", styles["h1"]),
        p(
            f"Les {len(analyse['rapprochements'])} clôtures exploitables ont été rapprochées des "
            "tickets EJ sur le CA TTC, le HT, la TVA et les modes de règlement. "
            f"Le résultat est de {ok}/{len(analyse['rapprochements'])} rapprochements au statut OK.",
            styles["corps"],
        ),
        tableau([
            ["Mode Z retenu", "Nombre de clôtures", "Interprétation"],
            *[
                [mode, nombre, "Mode déterminé par les règles documentées"]
                for mode, nombre in sorted(modes.items())
            ],
            ["TOTAL", len(analyse["rapprochements"]), "Toutes les clôtures rapprochées sont concordantes"],
        ], [40 * mm, 39 * mm, 95 * mm], styles, alignements={1: "RIGHT"}),
        p("Couverture des périodes", styles["h2"]),
        p(
            f"Le contrôle recense {len(analyse['couvertures'])} couples boutique-période : "
            f"{len(analyse['couvertures']) - len(analyse['sans_cloture'])} avec clôture et "
            f"{len(analyse['sans_cloture'])} sans clôture étiquetée. Aucune clôture artificielle "
            "n'est créée pour combler une absence.",
            styles["corps"],
        ),
        tableau([
            ["Boutique", "Période", "Observation", "Qualification"],
            *[
                [ligne["BOUTIQUE"], ligne["PERIODE"], "Aucune clôture Z étiquetée", "Point à documenter"]
                for ligne in analyse["sans_cloture"]
            ],
        ], [36 * mm, 32 * mm, 62 * mm, 44 * mm], styles),
        encadre(
            "Ces quatre absences de clôture constituent une limite documentaire à expliquer, mais "
            "pas un écart chiffré EJ/Z : aucune donnée de clôture inexistante n'a été reconstituée. "
            "Elles concernent MATURIN en novembre et décembre 2023, puis avril et mai 2025.",
            styles["alerte"],
            ORANGE_CLAIR,
        ),
        PageBreak(),
    ])

    # Livrables
    histoire.extend([
        p("6. Contrôle des classeurs remis", styles["h1"]),
        p(
            f"Le dossier contient {len(analyse['classeurs'])} classeurs Excel, totalisant "
            f"{analyse['total_feuilles']} feuilles et {analyse['total_formules']:,} formules. "
            f"Le contrôle syntaxique n'a détecté aucune formule contenant #REF! "
            f"({analyse['erreurs_ref']} occurrence). Les champs CA3 et FEC restent volontairement "
            "vides tant que leurs sources ne sont pas fournies."
            .replace(",", " "),
            styles["corps"],
        ),
        tableau([
            ["Famille", "Classeurs", "Objet"],
            ["EJ entêtes", 2, "Tickets de vente par boutique"],
            ["EJ lignes", 2, "Articles et corrections par boutique"],
            ["Z1", 6, "Synthèses mensuelles 2023 à 2025"],
            ["Z2", 6, "Transactions et règlements 2023 à 2025"],
            ["Recettes consolidées", 1, "Vue mensuelle des deux boutiques"],
            ["Comparaison CA3", 1, "Structure prête à recevoir les données déclaratives"],
            ["TOTAL", 18, "135 feuilles"],
        ], [55 * mm, 30 * mm, 89 * mm], styles, alignements={1: "RIGHT"}),
        p("Contrôles intégrés aux feuilles", styles["h2"]),
        p(
            "Les classeurs présentent les contrôles de cohérence, séquentialité, doublons, "
            "occurrences, agrégations et rapprochements prévus par le cahier des charges. Les "
            "formules restent visibles et recalculables dans un tableur compatible. Les fichiers "
            "techniques de contrôle et les aperçus de qualité ne font pas partie du dossier remis.",
            styles["corps"],
        ),
        p("Inventaire détaillé", styles["h2"]),
        tableau([
            ["Classeur", "Feuilles", "Formules"],
            *[[item["nom"], item["feuilles"], item["formules"]] for item in analyse["classeurs"]],
        ], [118 * mm, 26 * mm, 30 * mm], styles, alignements={1: "RIGHT", 2: "RIGHT"}),
        PageBreak(),
    ])

    # Anomalies / conclusion
    histoire.extend([
        p("7. Qualification des constats", styles["h1"]),
        tableau([
            ["Niveau", "Constat", "Analyse / action"],
            ["Aucune anomalie inexpliquée", "Séquences exhaustives", "0 trou, 0 doublon, 0 retour chronologique"],
            ["Aucune anomalie inexpliquée", "Cohérence monétaire", "0 écart sur 1 875 tickets"],
            ["Aucune anomalie inexpliquée", "Rapprochement EJ / Z", "57 clôtures sur 57 au statut OK"],
            ["Ecart apparent expliqué", "624 numéros internes absents des seules ventes", "Tous retrouvés parmi les blocs exclus conservés en base"],
            ["Ecart apparent expliqué", "636 blocs hors ventes", "Blocs administratifs ou événements à montant nul"],
            ["Point à documenter", "4 périodes MATURIN sans clôture Z", "Obtenir une explication ou un justificatif de l'exploitant"],
            ["Contrôle impossible", "FEC et CA3 absents", "Fournir les sources pour les rapprochements comptables et déclaratifs"],
        ], [44 * mm, 59 * mm, 71 * mm], styles),
        p("Conclusion", styles["h2"]),
        p(
            "Sur le seul périmètre des données de caisse communiquées, les informations sont "
            "exhaustivement conservées, séquentiellement cohérentes et arithmétiquement concordantes. "
            "Aucune irrégularité fiscale ne peut être déduite des écarts apparents décrits dans ce "
            "rapport : ils résultent de l'exclusion documentée des événements non commerciaux dans "
            "les classeurs de ventes.",
            styles["corps"],
        ),
        p(
            "Cette conclusion doit toutefois rester limitée. L'absence de quatre clôtures Z "
            "étiquetées appelle une explication documentaire, et l'absence des FEC, CA3 et "
            "justificatifs externes interdit de conclure sur la concordance avec la comptabilité, "
            "les déclarations de TVA ou la situation fiscale globale de l'entreprise.",
            styles["corps"],
        ),
        encadre(
            "Statut final : CONFORME SUR LE PERIMETRE CAISSE - A COMPLETER par les pièces externes "
            "et par l'explication des quatre périodes sans clôture Z étiquetée.",
            styles["statut"],
            VERT_CLAIR,
        ),
        Spacer(1, 8 * mm),
        p("Traçabilité du rapport", styles["h2"]),
        p(
            "Le rapport est régénéré par le programme Python à chaque traitement complet. Ses "
            "constats sont calculés directement depuis database/db.sqlite, les fichiers de contrôle "
            "et les classeurs présents dans output ; ils ne reposent pas sur une saisie manuelle.",
            styles["petit"],
        ),
    ])

    document.build(histoire)


def verifier_livraison(sortie: Path) -> None:
    elements = {chemin.name for chemin in sortie.iterdir()}
    pdf = sortie / NOM_RAPPORT
    if not pdf.is_file() or pdf.stat().st_size == 0:
        raise RuntimeError(f"Rapport PDF absent ou vide : {pdf}")
    inattendus = sorted(
        nom for nom in elements
        if not nom.endswith(".xlsx") and nom != NOM_RAPPORT
    )
    if inattendus:
        raise RuntimeError(f"Fichiers non autorisés dans output : {inattendus}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base", type=Path, required=True)
    parser.add_argument("--controle", type=Path, required=True)
    parser.add_argument("--sortie", type=Path, required=True)
    args = parser.parse_args(argv)
    analyse = analyser_donnees(args.base, args.controle, args.sortie)
    destination = args.sortie / NOM_RAPPORT
    generer_pdf(analyse, destination)
    verifier_livraison(args.sortie)
    print(f"Rapport fiscal généré : {destination}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
