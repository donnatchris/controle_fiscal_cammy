"""Construit en Python les 18 classeurs contractuels du dossier fiscal 751."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import subprocess
import tempfile

from collections import Counter
from copy import copy
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from shared.constantes import (
    ALIAS_COURT,
    NOM_COMPLET,
    ClasseurResolu751,
    SEPARATEUR_CSV,
    resoudre_classeur_751,
)


BOUTIQUES = ("MASSENA", "MATURIN")
ANNEES = (2023, 2024, 2025)
MOIS_PAR_ANNEE = {
    2023: [f"2023-{mois:02d}" for mois in range(1, 13)],
    2024: [f"2024-{mois:02d}" for mois in range(1, 13)],
    2025: [f"2025-{mois:02d}" for mois in range(1, 9)],
}

COLONNES_EJ = [
    "nomfichier", "E_NUM_INTERNE", "E_NUM_TICKET", "E_DATE_TICKET",
    "E_HEURE_TICKET", "E_HT1", "E_HT2", "E_HT3", "E_HT4", "E_TVA1",
    "E_TVA2", "E_TVA3", "E_TVA4", "E_HT_NON_TAXABLE", "E_TTC",
    "E_MDP_CB", "E_MDP_ESPECES", "E_MDP_CHEQUES",
]
COLONNES_LIGNES_EJ = [
    *COLONNES_EJ, "D_QUANTITE_ARTICLE", "D_LIBELLE_ARTICLE",
    "D_TAUX_TVA_ARTICLE", "D_MONTANT_ARTICLE", "D_CORRECTION", "D_AUTRE_INFO",
]
COLONNES_Z = [
    "nomfichier", "E_MODELE", "E_MACHINE", "E_RAPPORT", "E_FICHIER",
    "E_MODE", "E_COMPTEUR_Z", "E_DATE", "E_HEURE", "D_ENREGISTREMENT",
    "D_DESIGNATION", "D_QUANTITE", "D_MONTANT",
]
CHAMPS_NUMERIQUES = {
    "E_HT1", "E_HT2", "E_HT3", "E_HT4", "E_TVA1", "E_TVA2", "E_TVA3",
    "E_TVA4", "E_HT_NON_TAXABLE", "E_TTC", "E_MDP_CB", "E_MDP_ESPECES",
    "E_MDP_CHEQUES", "D_QUANTITE_ARTICLE", "D_MONTANT_ARTICLE",
    "D_CORRECTION", "D_QUANTITE", "D_MONTANT",
}
CIBLES_Z2 = ("CARTES", "CHEQUES", "CORRECTION", "ESPECES", "REF./TIROIR")
CIBLES_Z1 = (
    "CA BRUT", "CA NET", "CB.TIROIR", "CHQ.TIROIR", "ESP.TIROIR",
    "HORS TAXE 1", "TVA 1",
)

REMPLISSAGE_ENTETE = PatternFill("solid", fgColor="1F4E78")
REMPLISSAGE_SAISIE = PatternFill("solid", fgColor="FFF2CC")
REMPLISSAGE_ALERTE = PatternFill("solid", fgColor="FEF3C7")
REMPLISSAGE_ECART = PatternFill("solid", fgColor="FECACA")
POLICE_ALERTE = Font(name="Aptos", size=9, bold=True, color="92400E")
POLICE_ECART = Font(name="Aptos", size=9, bold=True, color="991B1B")
BORD_FIN = Side(style="thin", color="E2E8F0")
BORD_ENTETE = Side(style="medium", color="1F4E78")
FORMAT_MONTANT = "#,##0.00;[Red]-#,##0.00"


def valeur_decimal(valeur: Any) -> Decimal | None:
    if valeur in (None, ""):
        return None
    texte = str(valeur).replace(" ", "").replace(",", ".")
    try:
        return Decimal(texte)
    except Exception:
        return None


def nombre(valeur: Decimal | None) -> Decimal:
    return valeur if valeur is not None else Decimal("0")


def lire_csv(chemin: Path) -> list[dict[str, Any]]:
    with chemin.open(encoding="utf-8-sig", newline="") as fichier:
        lignes = []
        for ligne in csv.DictReader(fichier, delimiter=SEPARATEUR_CSV):
            lignes.append({
                champ: valeur_decimal(valeur) if champ in CHAMPS_NUMERIQUES else (valeur or "")
                for champ, valeur in ligne.items()
            })
    return lignes


def convertir_date(valeur: Any) -> Any:
    if not valeur or isinstance(valeur, datetime):
        return valeur or None
    texte = str(valeur).strip()
    for format_date in ("%Y%m%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(texte[:10], format_date)
        except ValueError:
            continue
    return valeur


def normaliser_cellule(entete: str, valeur: Any) -> Any:
    if entete in {"E_DATE_TICKET", "E_DATE"}:
        return convertir_date(valeur)
    return None if valeur is None else valeur


def nettoyer_nom(valeur: str) -> str:
    resultat = re.sub(r"[^A-Za-z0-9_]", "_", str(valeur))
    return f"_{resultat}" if resultat and not re.match(r"[A-Za-z_]", resultat[0]) else resultat


def periodes_reference(nom_fichier: str) -> list[str]:
    motifs = re.findall(r"(?:^|_)(0[1-9]|1[0-2])(20\d{2})(?=_|\.|$)", str(nom_fichier))
    return list(dict.fromkeys(f"{annee}-{mois}" for mois, annee in motifs))


def etiquette_periodes(nom_fichier: str) -> str:
    return "|".join(periodes_reference(nom_fichier))


def trier_tickets(lignes: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    def cle(ligne: dict[str, Any]) -> tuple[int, int, str]:
        return (
            int(ligne.get("E_NUM_INTERNE") or 0),
            int(ligne.get("E_NUM_TICKET") or 0),
            str(ligne.get("E_DATE_TICKET") or ""),
        )
    return sorted(lignes, key=cle)


def lignes_mensuelles(annee: int | None = None) -> list[dict[str, Any]]:
    mois = MOIS_PAR_ANNEE[annee] if annee else [m for a in ANNEES for m in MOIS_PAR_ANNEE[a]]
    return [{"AJ_ANNEE": int(m[:4]), "AJ_MOIS": m} for m in mois]


def largeur_colonne(nom: str) -> float:
    if nom == "nomfichier" or "FICHIER" in nom:
        return 25
    if any(terme in nom for terme in ("LIBELLE", "DESIGNATION", "COMMENTAIRE")):
        return 24
    if nom == "AJ_MOIS_Z":
        return 30
    if any(terme in nom for terme in ("DATE", "MOIS", "ANNEE")):
        return 14
    if any(terme in nom for terme in ("STATUT", "OBSERVATION")):
        return 18
    return 14


def format_colonne(nom: str) -> str | None:
    if nom in {"E_DATE_TICKET", "E_DATE"}:
        return "yyyy-mm-dd"
    if "ANNEE" in nom:
        return "0"
    if nom in {"E_NUM_INTERNE", "E_NUM_TICKET", "E_COMPTEUR_Z", "D_ENREGISTREMENT"}:
        return "@"
    if any(terme in nom for terme in ("QUANTITE", "OCCURRENCE", "COMPTE")):
        return "0"
    if (
        re.match(r"^(E_HT|E_TVA|E_TTC|E_MDP|D_MONTANT|D_CORRECTION|AJ_|SOMME_|MASSENA_|MATURIN_|MTT_)", nom)
        or "MONTANT" in nom
        or any(terme in nom for terme in ("TOTAL_HT", "TOTAL_TVA", "TOTAL_TTC"))
    ):
        return FORMAT_MONTANT
    return None


def ajouter_feuille(
    classeur: Workbook,
    nom: str,
    entetes: list[str],
    lignes: list[dict[str, Any]],
    nom_tableau: str,
) -> Any:
    if len(nom) > 31:
        raise ValueError(f"Nom de feuille XLSX trop long ({len(nom)} caractères) : {nom}")
    if nom in classeur.sheetnames:
        raise ValueError(f"Nom de feuille XLSX dupliqué : {nom}")
    feuille = classeur.create_sheet(nom)
    feuille.sheet_view.showGridLines = False
    feuille.freeze_panes = "A2"
    feuille.append(entetes)
    for ligne in lignes:
        feuille.append([normaliser_cellule(entete, ligne.get(entete)) for entete in entetes])

    for cellule in feuille[1]:
        cellule.fill = copy(REMPLISSAGE_ENTETE)
        cellule.font = Font(name="Aptos Display", size=10, bold=True, color="FFFFFF")
        cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cellule.border = Border(top=BORD_ENTETE, bottom=BORD_ENTETE, left=BORD_ENTETE, right=BORD_ENTETE)
    feuille.row_dimensions[1].height = 34

    derniere_ligne = max(len(lignes) + 1, 2)
    for index, entete in enumerate(entetes, start=1):
        lettre = get_column_letter(index)
        feuille.column_dimensions[lettre].width = largeur_colonne(entete)
        format_nombre = format_colonne(entete)
        for ligne in range(2, derniere_ligne + 1):
            cellule = feuille.cell(ligne, index)
            cellule.font = Font(name="Aptos", size=9)
            cellule.alignment = Alignment(vertical="center")
            cellule.border = Border(bottom=BORD_FIN)
            if format_nombre:
                cellule.number_format = format_nombre

    if lignes:
        reference = f"A1:{get_column_letter(len(entetes))}{len(lignes) + 1}"
        tableau = Table(displayName=nettoyer_nom(nom_tableau), ref=reference)
        tableau.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        feuille.add_table(tableau)
        feuille.auto_filter.ref = reference
    return feuille


def colonne_formules(
    feuille: Any,
    colonne: int,
    nombre_lignes: int,
    formule: Callable[[int, int], str],
) -> None:
    for index in range(nombre_lignes):
        feuille.cell(index + 2, colonne + 1, formule(index + 2, index))


def mise_en_forme_ecart(feuille: Any, colonne: int, nombre_lignes: int, tolerance: Decimal = Decimal("0.02")) -> None:
    if not nombre_lignes:
        return
    lettre = get_column_letter(colonne + 1)
    plage = f"{lettre}2:{lettre}{nombre_lignes + 1}"
    feuille.conditional_formatting.add(
        plage,
        CellIsRule(operator="greaterThan", formula=[str(tolerance)], fill=REMPLISSAGE_ECART, font=POLICE_ECART),
    )
    feuille.conditional_formatting.add(
        plage,
        CellIsRule(operator="lessThan", formula=[str(-tolerance)], fill=REMPLISSAGE_ECART, font=POLICE_ECART),
    )


def lire_lignes_feuille(
    feuille: Any,
    entetes: Iterable[str] | None = None,
) -> list[dict[str, Any]]:
    """Lit les valeurs d'une feuille déjà construite, source contractuelle immédiate."""
    entetes_observees = [cellule.value for cellule in feuille[1]]
    if not entetes_observees or any(not isinstance(entete, str) for entete in entetes_observees):
        raise ValueError(f"{feuille.title}: ligne d'entête invalide")
    selection = list(entetes) if entetes is not None else list(entetes_observees)
    absentes = [entete for entete in selection if entete not in entetes_observees]
    if absentes:
        raise ValueError(f"{feuille.title}: colonnes sources absentes : {absentes}")
    positions = {entete: entetes_observees.index(entete) + 1 for entete in selection}
    lignes: list[dict[str, Any]] = []
    for numero_ligne in range(2, feuille.max_row + 1):
        ligne = {
            entete: feuille.cell(numero_ligne, position).value
            for entete, position in positions.items()
        }
        if any(valeur not in (None, "") for valeur in ligne.values()):
            lignes.append(ligne)
    return lignes


def feuille_copie_valeurs(
    classeur: Workbook,
    nom: str,
    source: Any,
    entetes: list[str],
    tableau: str,
) -> Any:
    """Copie en valeur une feuille source, conformément aux mentions du CDC."""
    return ajouter_feuille(
        classeur,
        nom,
        entetes,
        lire_lignes_feuille(source, entetes),
        tableau,
    )


def nouveau_classeur() -> Workbook:
    classeur = Workbook()
    classeur.remove(classeur.active)
    classeur.calculation.fullCalcOnLoad = True
    classeur.calculation.forceFullCalc = True
    classeur.calculation.calcMode = "auto"
    return classeur


def totaux_ej_mensuels(lignes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resultat = {
        mois: {
            "AJ_ANNEE": int(mois[:4]), "AJ_MOIS": mois, "E_TTC": Decimal("0"),
            "E_MDP_CB": Decimal("0"), "E_MDP_CHEQUES": Decimal("0"),
            "E_MDP_ESPECES": Decimal("0"), "AJ_TOTAL_HT": Decimal("0"),
            "AJ_TOTAL_TVA_20": Decimal("0"),
        }
        for annee in ANNEES for mois in MOIS_PAR_ANNEE[annee]
    }
    for ligne in lignes:
        date = str(ligne.get("E_DATE_TICKET") or "")
        mois = f"{date[:4]}-{date[4:6]}" if re.fullmatch(r"\d{8}", date) else date[:7]
        if mois not in resultat:
            continue
        courant = resultat[mois]
        for champ in ("E_TTC", "E_MDP_CB", "E_MDP_CHEQUES", "E_MDP_ESPECES"):
            courant[champ] += nombre(ligne.get(champ))
        courant["AJ_TOTAL_HT"] += sum(
            (nombre(ligne.get(champ)) for champ in ("E_HT1", "E_HT2", "E_HT3", "E_HT4", "E_HT_NON_TAXABLE")),
            Decimal("0"),
        )
        courant["AJ_TOTAL_TVA_20"] += nombre(ligne.get("E_TVA1"))
    return list(resultat.values())


def decimal_cellule(valeur: Any) -> Decimal:
    return nombre(valeur_decimal(valeur))


def mois_ticket(valeur: Any) -> str:
    if isinstance(valeur, datetime):
        return valeur.strftime("%Y-%m")
    texte = str(valeur or "")
    if re.fullmatch(r"\d{8}", texte):
        return f"{texte[:4]}-{texte[4:6]}"
    return texte[:7]


def agreger_mois_ej(
    lignes: Iterable[dict[str, Any]],
    correspondances: dict[str, str],
) -> list[dict[str, Any]]:
    resultat = {
        mois["AJ_MOIS"]: {**mois, **{cible: Decimal("0") for cible in correspondances}}
        for mois in lignes_mensuelles()
    }
    for ligne in lignes:
        mois = str(ligne.get("AJ_MOIS") or mois_ticket(ligne.get("E_DATE_TICKET")))
        if mois not in resultat:
            continue
        for cible, source in correspondances.items():
            resultat[mois][cible] += decimal_cellule(ligne.get(source))
    return list(resultat.values())


def valeurs_entetes_completes_pour_agregation(
    feuille: Any,
) -> list[dict[str, Any]]:
    """Matérialise les deux formules de la feuille complète pour ses agrégats enfants."""
    lignes = lire_lignes_feuille(feuille)
    for ligne in lignes:
        ligne["AJ_TOTAL_HT"] = sum(
            (
                decimal_cellule(ligne.get(champ))
                for champ in ("E_HT1", "E_HT2", "E_HT3", "E_HT4", "E_HT_NON_TAXABLE")
            ),
            start=Decimal("0"),
        )
        ligne["AJ_TOTAL_TVA_20"] = decimal_cellule(ligne.get("E_TVA1"))
    return lignes


def sommer_periode(
    lignes: Iterable[dict[str, Any]],
    periode: str,
    champ: str,
) -> Decimal:
    mois = set(str(periode).split("|"))
    return sum(
        (
            decimal_cellule(ligne.get(champ))
            for ligne in lignes
            if str(ligne.get("AJ_MOIS") or ligne.get("AJ_MOIS_Z")) in mois
        ),
        start=Decimal("0"),
    )


class Constructeur751:
    def __init__(
        self,
        staging: Path,
        controle: Path,
        sortie: Path,
        qa_dir: Path,
        qa: bool = False,
    ) -> None:
        self.csv = staging
        self.controle = controle
        self.xlsx = sortie
        self.qa = qa
        self.qa_dir = qa_dir
        self.xlsx.mkdir(parents=True, exist_ok=True)
        self.filiations_actives: list[dict[str, Any]] = []
        self.donnees: dict[str, dict[str, Any]] = {"entetes": {}, "lignes": {}, "z1": {}, "z2": {}}
        for boutique in BOUTIQUES:
            self.donnees["entetes"][boutique] = lire_csv(self.csv / f"EJ_ENTETES_TICKETS_{boutique}.csv")
            self.donnees["lignes"][boutique] = lire_csv(self.csv / f"EJ_LIGNES_TICKETS_{boutique}.csv")
            self.donnees["z1"][boutique] = {}
            self.donnees["z2"][boutique] = {}
            for annee in ANNEES:
                self.donnees["z1"][boutique][annee] = lire_csv(self.csv / f"Z1_SyntheseMois_TOUS_{annee}_{boutique}.csv")
                self.donnees["z2"][boutique][annee] = lire_csv(self.csv / f"Z2_TransactionsMois_TOUS_{annee}_{boutique}.csv")

        self.manifeste: dict[str, Any] = {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "generator": "Python/openpyxl",
            "excelDirectory": self.xlsx.name,
            "workbooks": [],
            "csvFiles": [
                nom
                for boutique in BOUTIQUES
                for nom in (
                    f"EJ_ENTETES_TICKETS_{boutique}.csv",
                    f"EJ_LIGNES_TICKETS_{boutique}.csv",
                    *(f"Z{niveau}_{'SyntheseMois' if niveau == 1 else 'TransactionsMois'}_TOUS_{annee}_{boutique}.csv"
                      for annee in ANNEES for niveau in (1, 2)),
                )
            ],
            "checks": [],
        }

    def commencer_filiations(self) -> None:
        self.filiations_actives = []

    def noter_filiation(
        self,
        cible: str,
        sources: Iterable[str],
        operation: str,
    ) -> None:
        self.filiations_actives.append({
            "targetSheet": cible,
            "immediateSources": list(sources),
            "operation": operation,
        })

    def lire_feuille_produite(
        self,
        nom_classeur: str,
        nom_feuille: str,
        entetes: Iterable[str] | None = None,
    ) -> list[dict[str, Any]]:
        chemin = self.xlsx / nom_classeur
        if not chemin.is_file():
            raise FileNotFoundError(f"Classeur source non produit : {chemin}")
        classeur = load_workbook(chemin, read_only=True, data_only=True)
        try:
            if nom_feuille not in classeur.sheetnames:
                raise KeyError(f"{nom_classeur}: feuille source absente : {nom_feuille}")
            return lire_lignes_feuille(classeur[nom_feuille], entetes)
        finally:
            classeur.close()

    def periodes_cloture(self, annee: int, feuille_complete: Any) -> list[dict[str, Any]]:
        lignes = lire_lignes_feuille(feuille_complete, ["AJ_MOIS_Z"])
        composites = list(dict.fromkeys(
            str(ligne["AJ_MOIS_Z"])
            for ligne in lignes
            if "|" in str(ligne["AJ_MOIS_Z"])
        ))
        composantes = {mois for etiquette in composites for mois in etiquette.split("|")}
        etiquettes = [mois for mois in MOIS_PAR_ANNEE[annee] if mois not in composantes] + composites
        etiquettes.sort(key=lambda valeur: valeur.split("|")[-1])
        return [{"AJ_ANNEE_Z": annee, "AJ_MOIS_Z": etiquette} for etiquette in etiquettes]

    def exporter(self, classeur: Workbook, definition: ClasseurResolu751) -> None:
        nom_fichier = definition.nom_fichier
        feuilles_demandees = list(definition.noms_feuilles(NOM_COMPLET))
        alias_courts = list(definition.noms_feuilles(ALIAS_COURT))
        noms = classeur.sheetnames
        if len(noms) != len(feuilles_demandees):
            raise RuntimeError(f"{nom_fichier}: {len(noms)} feuilles au lieu de {len(feuilles_demandees)}")
        if noms != alias_courts:
            raise RuntimeError(
                f"{nom_fichier}: ordre ou alias de feuilles invalide : "
                f"observes={noms}, attendus={alias_courts}"
            )
        if hasattr(self, "filiations_actives"):
            cibles_filiation = [entree["targetSheet"] for entree in self.filiations_actives]
            if cibles_filiation != noms:
                raise RuntimeError(
                    f"{nom_fichier}: filiation incomplète ou désordonnée : "
                    f"observee={cibles_filiation}, attendue={noms}"
                )
            for entree in self.filiations_actives:
                if not entree["immediateSources"]:
                    raise RuntimeError(
                        f"{nom_fichier}: aucune source immédiate pour {entree['targetSheet']}"
                    )
        erreurs = [
            f"{feuille.title}!{cellule.coordinate}={cellule.value}"
            for feuille in classeur.worksheets
            for ligne in feuille.iter_rows()
            for cellule in ligne
            if isinstance(cellule.value, str)
            and cellule.value.startswith("=")
            and any(erreur in cellule.value.upper() for erreur in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"))
        ]
        if erreurs:
            raise RuntimeError(f"{nom_fichier}: formules invalides: {erreurs[:5]}")
        destination = self.xlsx / nom_fichier
        classeur.save(destination)
        verification = load_workbook(destination, read_only=True, data_only=False)
        try:
            if verification.sheetnames != noms:
                raise RuntimeError(f"{nom_fichier}: feuilles altérées après sauvegarde")
        finally:
            verification.close()
        self.manifeste["workbooks"].append({
            "fileName": nom_fichier,
            "sheetCount": len(noms),
            "sheets": noms,
            "requestedSheets": feuilles_demandees,
            "sheetNameMode": ALIAS_COURT,
            "sheetNameMappings": [
                {
                    "nomComplet": feuille.nom_complet,
                    "aliasCourt": feuille.alias_court,
                    "nomProduit": feuille.alias_court,
                }
                for feuille in definition.feuilles
            ],
            "sheetLineage": list(getattr(self, "filiations_actives", [])),
        })
        self.manifeste["checks"].append({"fileName": nom_fichier, "formulaErrors": 0})

    def construire_entetes(self, boutique: str) -> None:
        classeur = nouveau_classeur()
        self.commencer_filiations()
        definition = resoudre_classeur_751("ej_entetes", boutique=boutique)
        noms = list(definition.noms_feuilles(ALIAS_COURT))
        source_csv = f"EJ_ENTETES_TICKETS_{boutique}.csv"
        brut = ajouter_feuille(
            classeur, noms[0], COLONNES_EJ,
            self.donnees["entetes"][boutique], "Entetes0",
        )
        self.noter_filiation(noms[0], [source_csv], "ingestion")
        triees = trier_tickets(lire_lignes_feuille(brut, COLONNES_EJ))
        tri = ajouter_feuille(classeur, noms[1], COLONNES_EJ, triees, "EntetesTri")
        self.noter_filiation(noms[1], [noms[0]], "copie et tri croissant E_NUM_INTERNE")

        entetes_ctrl = [*COLONNES_EJ, "AJ_TVA1_CALCULE", "AJ_ECART_TVA1", "AJ_TTC_CALCULE", "AJ_ECART_TTC", "AJ_SOLDE_DU", "AJ_SOLDE_TOUS_MDP"]
        lignes_ctrl = lire_lignes_feuille(tri, COLONNES_EJ)
        ctrl = ajouter_feuille(classeur, noms[2], entetes_ctrl, lignes_ctrl, "EntetesCtrl")
        self.noter_filiation(noms[2], [noms[1]], "copie et enrichissement de cohérence")
        colonnes = {nom: get_column_letter(index + 1) for index, nom in enumerate(entetes_ctrl)}
        formules = {
            "AJ_TVA1_CALCULE": lambda r: f"={colonnes['E_HT1']}{r}*20%",
            "AJ_ECART_TVA1": lambda r: f"={colonnes['E_TVA1']}{r}-{colonnes['AJ_TVA1_CALCULE']}{r}",
            "AJ_TTC_CALCULE": lambda r: f"={colonnes['E_HT1']}{r}+{colonnes['E_TVA1']}{r}",
            "AJ_ECART_TTC": lambda r: f"={colonnes['E_TTC']}{r}-{colonnes['AJ_TTC_CALCULE']}{r}",
            "AJ_SOLDE_DU": lambda r: f"={colonnes['E_TTC']}{r}-({colonnes['E_MDP_CB']}{r}+{colonnes['E_MDP_CHEQUES']}{r})",
            "AJ_SOLDE_TOUS_MDP": lambda r: f"={colonnes['E_TTC']}{r}-({colonnes['E_MDP_CB']}{r}+{colonnes['E_MDP_ESPECES']}{r}+{colonnes['E_MDP_CHEQUES']}{r})",
        }
        for nom, formule in formules.items():
            colonne_formules(ctrl, entetes_ctrl.index(nom), len(lignes_ctrl), lambda r, _i, f=formule: f(r))
        for nom in ("AJ_ECART_TVA1", "AJ_ECART_TTC", "AJ_SOLDE_TOUS_MDP"):
            mise_en_forme_ecart(ctrl, entetes_ctrl.index(nom), len(lignes_ctrl))

        entetes_seq = ["nomfichier", "E_NUM_INTERNE", "E_NUM_TICKET", "E_DATE_TICKET", "E_HEURE_TICKET", "AJ_TROU_NUM_TICKET"]
        lignes_seq = lire_lignes_feuille(ctrl, entetes_seq[:5])
        seq = ajouter_feuille(classeur, noms[3], entetes_seq, lignes_seq, "EntetesSeq")
        self.noter_filiation(noms[3], [noms[2]], "copie en valeur et analyse séquentielle")
        colonne_formules(seq, 5, len(lignes_seq), lambda r, i: '=""' if i == 0 else f"=C{r}-C{r-1}")

        source_seq = lire_lignes_feuille(seq, ["E_NUM_INTERNE", "E_NUM_TICKET"])
        comptes_internes = Counter(ligne["E_NUM_INTERNE"] for ligne in source_seq if ligne.get("E_NUM_INTERNE") not in (None, ""))
        valeurs_internes = [
            {"E_NUM_INTERNE": valeur, "COMPTER_E_NUM_INTERNE": comptes_internes[valeur]}
            for valeur in sorted(comptes_internes, key=int)
        ]
        entetes_occ_int = ["E_NUM_INTERNE", "COMPTER_E_NUM_INTERNE"]
        occ_int = ajouter_feuille(classeur, noms[4], entetes_occ_int, valeurs_internes, "OccInt")
        self.noter_filiation(noms[4], [noms[3]], "agrégation des occurrences")
        feuille_copie_valeurs(classeur, noms[5], occ_int, entetes_occ_int, "DupInt")
        self.noter_filiation(noms[5], [noms[4]], "copie en valeur")

        comptes_tickets = Counter(ligne["E_NUM_TICKET"] for ligne in source_seq if ligne.get("E_NUM_TICKET") not in (None, ""))
        valeurs_tickets = [
            {"E_NUM_TICKET": valeur, "COMPTER_E_NUM_TICKET": comptes_tickets[valeur]}
            for valeur in sorted(comptes_tickets, key=int)
        ]
        entetes_occ_ticket = ["E_NUM_TICKET", "COMPTER_E_NUM_TICKET"]
        occ_ticket = ajouter_feuille(classeur, noms[6], entetes_occ_ticket, valeurs_tickets, "OccTicket")
        self.noter_filiation(noms[6], [noms[3]], "agrégation des occurrences")
        feuille_copie_valeurs(classeur, noms[7], occ_ticket, entetes_occ_ticket, "DupTicket")
        self.noter_filiation(noms[7], [noms[6]], "copie en valeur")

        entetes_completes = [*COLONNES_EJ, "AJ_TOTAL_HT", "AJ_TOTAL_TVA_20", "AJ_ANNEE", "AJ_MOIS"]
        lignes_completes = []
        for ligne in lire_lignes_feuille(tri, COLONNES_EJ):
            mois = mois_ticket(ligne["E_DATE_TICKET"])
            lignes_completes.append({
                **ligne,
                "AJ_ANNEE": int(mois[:4]),
                "AJ_MOIS": mois,
            })
        complete = ajouter_feuille(classeur, noms[8], entetes_completes, lignes_completes, "EntetesCplte")
        self.noter_filiation(noms[8], [noms[1]], "copie, enrichissement année/mois et formules HT/TVA")
        colonnes_completes = {
            nom: get_column_letter(index + 1)
            for index, nom in enumerate(entetes_completes)
        }
        colonne_formules(
            complete,
            entetes_completes.index("AJ_TOTAL_HT"),
            len(lignes_completes),
            lambda r, _i: "=" + "+".join(
                f"{colonnes_completes[champ]}{r}"
                for champ in ("E_HT1", "E_HT2", "E_HT3", "E_HT4", "E_HT_NON_TAXABLE")
            ),
        )
        colonne_formules(
            complete,
            entetes_completes.index("AJ_TOTAL_TVA_20"),
            len(lignes_completes),
            lambda r, _i: f"={colonnes_completes['E_TVA1']}{r}",
        )
        lignes_complete_agregeables = valeurs_entetes_completes_pour_agregation(complete)

        entetes_enc = ["AJ_ANNEE", "AJ_MOIS", "SOMME_E_TTC", "SOMME_E_MDP_CB", "SOMME_E_MDP_CHEQUES", "SOMME_E_MDP_ESPECES"]
        lignes_enc = agreger_mois_ej(lignes_complete_agregeables, {
            "SOMME_E_TTC": "E_TTC",
            "SOMME_E_MDP_CB": "E_MDP_CB",
            "SOMME_E_MDP_CHEQUES": "E_MDP_CHEQUES",
            "SOMME_E_MDP_ESPECES": "E_MDP_ESPECES",
        })
        enc = ajouter_feuille(classeur, noms[9], entetes_enc, lignes_enc, "EncMensuelTD")
        self.noter_filiation(noms[9], [noms[8]], "agrégation mensuelle")
        feuille_copie_valeurs(classeur, noms[10], enc, entetes_enc, "EncMensuelCopie")
        self.noter_filiation(noms[10], [noms[9]], "copie en valeur")

        entetes_rec = ["AJ_ANNEE", "AJ_MOIS", "SOMME_AJ_TOTAL_HT", "SOMME_AJ_TOTAL_TVA_20", "SOMME_E_TTC"]
        lignes_rec = agreger_mois_ej(lignes_complete_agregeables, {
            "SOMME_AJ_TOTAL_HT": "AJ_TOTAL_HT",
            "SOMME_AJ_TOTAL_TVA_20": "AJ_TOTAL_TVA_20",
            "SOMME_E_TTC": "E_TTC",
        })
        rec = ajouter_feuille(classeur, noms[11], entetes_rec, lignes_rec, "RecMensuelTD")
        self.noter_filiation(noms[11], [noms[8]], "agrégation mensuelle")
        feuille_copie_valeurs(classeur, noms[12], rec, entetes_rec, "RecMensuelCopie")
        self.noter_filiation(noms[12], [noms[11]], "copie en valeur")
        self.exporter(classeur, definition)

    def construire_lignes(self, boutique: str) -> None:
        classeur = nouveau_classeur()
        self.commencer_filiations()
        definition = resoudre_classeur_751("ej_lignes", boutique=boutique)
        noms = list(definition.noms_feuilles(ALIAS_COURT))
        source_csv = f"EJ_LIGNES_TICKETS_{boutique}.csv"
        brut = ajouter_feuille(
            classeur, noms[0], COLONNES_LIGNES_EJ,
            self.donnees["lignes"][boutique], "Lignes0",
        )
        self.noter_filiation(noms[0], [source_csv], "ingestion")
        lignes_triees = trier_tickets(lire_lignes_feuille(brut, COLONNES_LIGNES_EJ))
        tri = ajouter_feuille(classeur, noms[1], COLONNES_LIGNES_EJ, lignes_triees, "LignesTri")
        self.noter_filiation(noms[1], [noms[0]], "copie et tri croissant E_NUM_INTERNE")
        ctrl = feuille_copie_valeurs(classeur, noms[2], tri, COLONNES_LIGNES_EJ, "LignesCtrl")
        self.noter_filiation(noms[2], [noms[1]], "copie en valeur")

        lignes_ctrl = lire_lignes_feuille(ctrl, COLONNES_LIGNES_EJ)
        groupes: dict[Any, list[dict[str, Any]]] = {}
        for ligne in lignes_ctrl:
            groupes.setdefault(ligne["E_NUM_TICKET"], []).append(ligne)
        totaux = []
        for numero, groupe in groupes.items():
            totaux.append({
                "E_NUM_TICKET": numero,
                "E_TTC": decimal_cellule(groupe[0].get("E_TTC")),
                "COMPTER_D_LIBELLE_ARTICLE": sum(1 for ligne in groupe if ligne.get("D_LIBELLE_ARTICLE") not in (None, "")),
                "SOMME_D_MONTANT_ARTICLE": sum((decimal_cellule(ligne.get("D_MONTANT_ARTICLE")) for ligne in groupe), start=Decimal("0")),
                "SOMME_D_CORRECTION": sum((decimal_cellule(ligne.get("D_CORRECTION")) for ligne in groupe), start=Decimal("0")),
            })
        entetes_totaux = ["E_NUM_TICKET", "E_TTC", "COMPTER_D_LIBELLE_ARTICLE", "SOMME_D_MONTANT_ARTICLE", "SOMME_D_CORRECTION"]
        feuille_totaux = ajouter_feuille(classeur, noms[3], entetes_totaux, totaux, "TotalLignes")
        self.noter_filiation(noms[3], [noms[2]], "agrégation par numéro de ticket")

        entetes_coherence = [*entetes_totaux, "AJ_ECART_TTC"]
        lignes_coherence = lire_lignes_feuille(feuille_totaux, entetes_totaux)
        coherence = ajouter_feuille(classeur, noms[4], entetes_coherence, lignes_coherence, "CoherenceEnteteLigne")
        self.noter_filiation(noms[4], [noms[3]], "copie en valeur et formule d'écart")
        colonnes_coherence = {
            nom: get_column_letter(index + 1)
            for index, nom in enumerate(entetes_coherence)
        }
        colonne_formules(
            coherence,
            entetes_coherence.index("AJ_ECART_TTC"),
            len(lignes_coherence),
            lambda r, _i: (
                f"={colonnes_coherence['E_TTC']}{r}-("
                f"{colonnes_coherence['SOMME_D_MONTANT_ARTICLE']}{r}+"
                f"{colonnes_coherence['SOMME_D_CORRECTION']}{r})"
            ),
        )
        mise_en_forme_ecart(coherence, 5, len(totaux))

        libelles = Counter(ligne["D_LIBELLE_ARTICLE"] for ligne in lignes_ctrl if ligne.get("D_LIBELLE_ARTICLE"))
        lignes_libelles = [{"D_LIBELLE_ARTICLE": valeur, "COMPTER_D_LIBELLE_ARTICLE": libelles[valeur]} for valeur in sorted(libelles)]
        entetes_libelles = ["D_LIBELLE_ARTICLE", "COMPTER_D_LIBELLE_ARTICLE"]
        ajouter_feuille(classeur, noms[5], entetes_libelles, lignes_libelles, "OccLibelle")
        self.noter_filiation(noms[5], [noms[2]], "agrégation des occurrences")
        taux = Counter(ligne["D_TAUX_TVA_ARTICLE"] for ligne in lignes_ctrl if ligne.get("D_TAUX_TVA_ARTICLE"))
        lignes_taux = [{"D_TAUX_TVA_ARTICLE": valeur, "COMPTER_D_TAUX_TVA_ARTICLE": taux[valeur]} for valeur in sorted(taux)]
        entetes_taux = ["D_TAUX_TVA_ARTICLE", "COMPTER_D_TAUX_TVA_ARTICLE"]
        ajouter_feuille(classeur, noms[6], entetes_taux, lignes_taux, "OccTxTva")
        self.noter_filiation(noms[6], [noms[2]], "agrégation des occurrences")
        self.exporter(classeur, definition)

    @staticmethod
    def entetes_larges(cibles: Iterable[str], quantite: bool) -> list[str]:
        resultat = ["AJ_ANNEE_Z", "AJ_MOIS_Z"]
        for cible in cibles:
            cle = nettoyer_nom(cible).upper()
            if quantite:
                resultat.append(f"{cle}_D_QUANTITE")
            resultat.append(f"{cle}_D_MONTANT")
        return resultat

    def ajouter_mode_large(
        self,
        classeur: Workbook,
        nom: str,
        mois: list[dict[str, Any]],
        mode: str,
        cibles: tuple[str, ...],
        feuille_td: Any,
        quantite: bool,
        tableau: str,
    ) -> tuple[Any, list[str]]:
        entetes = self.entetes_larges(cibles, quantite)
        source = lire_lignes_feuille(feuille_td)
        lignes = []
        for periode in mois:
            ligne = dict(periode)
            for cible in cibles:
                cle = nettoyer_nom(cible).upper()
                selection = [
                    source_ligne for source_ligne in source
                    if source_ligne["AJ_MOIS_Z"] == periode["AJ_MOIS_Z"]
                    and source_ligne["E_MODE"] == mode
                    and source_ligne["D_DESIGNATION"] == cible
                ]
                if quantite:
                    ligne[f"{cle}_D_QUANTITE"] = sum((decimal_cellule(item.get("D_QUANTITE")) for item in selection), start=Decimal("0"))
                ligne[f"{cle}_D_MONTANT"] = sum((decimal_cellule(item.get("D_MONTANT")) for item in selection), start=Decimal("0"))
            lignes.append(ligne)
        feuille = ajouter_feuille(classeur, nom, entetes, lignes, tableau)
        return feuille, entetes

    def construire_z2(self, boutique: str, annee: int) -> None:
        classeur = nouveau_classeur()
        self.commencer_filiations()
        definition = resoudre_classeur_751("z2", boutique=boutique, annee=annee)
        noms = list(definition.noms_feuilles(ALIAS_COURT))
        source_csv = f"Z2_TransactionsMois_TOUS_{annee}_{boutique}.csv"
        brut = ajouter_feuille(classeur, noms[0], COLONNES_Z, self.donnees["z2"][boutique][annee], "Z2Raw")
        self.noter_filiation(noms[0], [source_csv], "ingestion")
        entetes_completes = [*COLONNES_Z, "AJ_ANNEE_Z", "AJ_MOIS_Z"]
        completes = []
        for ligne in lire_lignes_feuille(brut, COLONNES_Z):
            periodes = periodes_reference(str(ligne["nomfichier"]))
            completes.append({
                **ligne,
                "AJ_ANNEE_Z": int(periodes[0][:4]),
                "AJ_MOIS_Z": "|".join(periodes),
            })
        complete = ajouter_feuille(classeur, noms[1], entetes_completes, completes, "Z2Cplte")
        self.noter_filiation(noms[1], [noms[0]], "copie en valeur et enrichissement année/mois Z")
        mois = self.periodes_cloture(annee, complete)
        source_complete = lire_lignes_feuille(complete)
        td_lignes = []
        for periode in mois:
            for mode in ("Z", "ZZ1", "ZZ2"):
                for designation in CIBLES_Z2:
                    selection = [
                        ligne for ligne in source_complete
                        if ligne["AJ_MOIS_Z"] == periode["AJ_MOIS_Z"]
                        and ligne["E_MODE"] == mode
                        and ligne["D_DESIGNATION"] == designation
                    ]
                    td_lignes.append({
                        **periode,
                        "E_MODE": mode,
                        "D_DESIGNATION": designation,
                        "D_QUANTITE": sum((decimal_cellule(ligne.get("D_QUANTITE")) for ligne in selection), start=Decimal("0")),
                        "D_MONTANT": sum((decimal_cellule(ligne.get("D_MONTANT")) for ligne in selection), start=Decimal("0")),
                    })
        td_entetes = ["AJ_ANNEE_Z", "AJ_MOIS_Z", "E_MODE", "D_DESIGNATION", "D_QUANTITE", "D_MONTANT"]
        td = ajouter_feuille(classeur, noms[2], td_entetes, td_lignes, "Z2TotalNature")
        self.noter_filiation(noms[2], [noms[1]], "agrégation par mois, mode et nature de transaction")
        entetes_modes: dict[str, list[str]] = {}
        feuilles_modes: dict[str, Any] = {}
        for index, mode in enumerate(("ZZ1", "ZZ2", "Z")):
            feuilles_modes[mode], entetes_modes[mode] = self.ajouter_mode_large(
                classeur, noms[3 + index], mois, mode, CIBLES_Z2, td, True, f"Z2Mode{mode}",
            )
            self.noter_filiation(noms[3 + index], [noms[2]], f"copie en valeur filtrée sur le mode {mode}")
        if boutique == "MASSENA":
            entetes_comp = ["AJ_ANNEE_Z", "AJ_MOIS_Z", *[element for cible in CIBLES_Z2 for element in (f"{nettoyer_nom(cible).upper()}_AJ_ECART_QTE", f"{nettoyer_nom(cible).upper()}_AJ_ECART_MONTANT")]]
            zz1 = {ligne["AJ_MOIS_Z"]: ligne for ligne in lire_lignes_feuille(feuilles_modes["ZZ1"])}
            zz2 = {ligne["AJ_MOIS_Z"]: ligne for ligne in lire_lignes_feuille(feuilles_modes["ZZ2"])}
            lignes_comp = []
            for periode in mois:
                etiquette = periode["AJ_MOIS_Z"]
                ligne = dict(periode)
                for cible in CIBLES_Z2:
                    cle = nettoyer_nom(cible).upper()
                    ligne[f"{cle}_AJ_ECART_QTE"] = decimal_cellule(zz1[etiquette][f"{cle}_D_QUANTITE"]) - decimal_cellule(zz2[etiquette][f"{cle}_D_QUANTITE"])
                    ligne[f"{cle}_AJ_ECART_MONTANT"] = decimal_cellule(zz1[etiquette][f"{cle}_D_MONTANT"]) - decimal_cellule(zz2[etiquette][f"{cle}_D_MONTANT"])
                lignes_comp.append(ligne)
            comp = ajouter_feuille(classeur, noms[6], entetes_comp, lignes_comp, "CompareZ2Modes")
            self.noter_filiation(noms[6], [noms[3], noms[4]], "comparaison des modes ZZ1 et ZZ2")
            for colonne in range(3, len(entetes_comp), 2):
                mise_en_forme_ecart(comp, colonne, len(mois))
        compare_nom = noms[-1]
        mode_retenu = "ZZ1" if boutique == "MASSENA" else "Z"
        entetes_compare = ["AJ_ANNEE_Z", "AJ_MOIS_Z", "CARTES_Z2", "CARTES_EJ", "AJ_ECART_CARTES", "CHEQUES_Z2", "CHEQUES_EJ", "AJ_ECART_CHEQUES", "ESPECES_Z2", "ESPECES_EJ", "AJ_ECART_ESPECES"]
        definition_ej = resoudre_classeur_751("ej_entetes", boutique=boutique)
        feuille_ej = definition_ej.noms_feuilles(ALIAS_COURT)[10]
        ej = self.lire_feuille_produite(definition_ej.nom_fichier, feuille_ej)
        modes = {ligne["AJ_MOIS_Z"]: ligne for ligne in lire_lignes_feuille(feuilles_modes[mode_retenu])}
        lignes_compare = []
        for periode in mois:
            etiquette = periode["AJ_MOIS_Z"]
            mode_ligne = modes[etiquette]
            valeurs = {
                "CARTES_Z2": decimal_cellule(mode_ligne["CARTES_D_MONTANT"]),
                "CARTES_EJ": sommer_periode(ej, etiquette, "SOMME_E_MDP_CB"),
                "CHEQUES_Z2": decimal_cellule(mode_ligne["CHEQUES_D_MONTANT"]),
                "CHEQUES_EJ": sommer_periode(ej, etiquette, "SOMME_E_MDP_CHEQUES"),
                "ESPECES_Z2": decimal_cellule(mode_ligne["ESPECES_D_MONTANT"]),
                "ESPECES_EJ": sommer_periode(ej, etiquette, "SOMME_E_MDP_ESPECES"),
            }
            lignes_compare.append({
                **periode, **valeurs,
                "AJ_ECART_CARTES": valeurs["CARTES_Z2"] - valeurs["CARTES_EJ"],
                "AJ_ECART_CHEQUES": valeurs["CHEQUES_Z2"] - valeurs["CHEQUES_EJ"],
                "AJ_ECART_ESPECES": valeurs["ESPECES_Z2"] - valeurs["ESPECES_EJ"],
            })
        compare = ajouter_feuille(classeur, compare_nom, entetes_compare, lignes_compare, "CompareZ2EJ")
        self.noter_filiation(compare_nom, [feuilles_modes[mode_retenu].title, f"{definition_ej.nom_fichier}#{feuille_ej}"], "comparaison Z2/EJ")
        for ecart in (4, 7, 10):
            mise_en_forme_ecart(compare, ecart, len(mois))
        self.exporter(classeur, definition)

    def construire_z1(self, boutique: str, annee: int) -> None:
        classeur = nouveau_classeur()
        self.commencer_filiations()
        mode_retenu = "ZZ1" if boutique == "MASSENA" else "Z"
        definition = resoudre_classeur_751("z1", boutique=boutique, annee=annee)
        noms = list(definition.noms_feuilles(ALIAS_COURT))
        source_csv = f"Z1_SyntheseMois_TOUS_{annee}_{boutique}.csv"
        brut = ajouter_feuille(classeur, noms[0], COLONNES_Z, self.donnees["z1"][boutique][annee], "Z1Raw")
        self.noter_filiation(noms[0], [source_csv], "ingestion")
        entetes_completes = [*COLONNES_Z, "AJ_ANNEE_Z", "AJ_MOIS_Z"]
        completes = []
        for ligne in lire_lignes_feuille(brut, COLONNES_Z):
            periodes = periodes_reference(str(ligne["nomfichier"]))
            completes.append({**ligne, "AJ_ANNEE_Z": int(periodes[0][:4]), "AJ_MOIS_Z": "|".join(periodes)})
        complete = ajouter_feuille(classeur, noms[1], entetes_completes, completes, "Z1Cplte")
        self.noter_filiation(noms[1], [noms[0]], "copie en valeur et enrichissement année/mois Z")
        source_complete = lire_lignes_feuille(complete)
        uniques = {}
        for ligne in source_complete:
            cle = (ligne["AJ_MOIS_Z"], ligne["E_DATE"], ligne["E_FICHIER"], ligne["E_MODE"])
            uniques.setdefault(cle, {"AJ_ANNEE_Z": ligne["AJ_ANNEE_Z"], "AJ_MOIS_Z": ligne["AJ_MOIS_Z"], "E_DATE": ligne["E_DATE"], "E_FICHIER": ligne["E_FICHIER"], "E_MODE": ligne["E_MODE"]})
        occurrences = sorted(uniques.values(), key=lambda ligne: (ligne["AJ_MOIS_Z"], ligne["E_MODE"], ligne["E_FICHIER"]))
        for ligne in occurrences:
            ligne["OCCURRENCES"] = sum(
                1 for source in source_complete
                if source["AJ_MOIS_Z"] == ligne["AJ_MOIS_Z"]
                and source["E_DATE"] == ligne["E_DATE"]
                and source["E_FICHIER"] == ligne["E_FICHIER"]
                and source["E_MODE"] == ligne["E_MODE"]
            )
        entetes_occ = ["AJ_ANNEE_Z", "AJ_MOIS_Z", "E_DATE", "E_FICHIER", "E_MODE", "OCCURRENCES"]
        ajouter_feuille(classeur, noms[2], entetes_occ, occurrences, "OccZ1FichMode")
        self.noter_filiation(noms[2], [noms[1]], "agrégation des occurrences fichier/mode")
        mois = self.periodes_cloture(annee, complete)
        td_lignes = []
        for periode in mois:
            for mode in ("Z", "ZZ1", "ZZ2"):
                for designation in CIBLES_Z1:
                    selection = [
                        ligne for ligne in source_complete
                        if ligne["AJ_MOIS_Z"] == periode["AJ_MOIS_Z"]
                        and ligne["E_MODE"] == mode
                        and ligne["D_DESIGNATION"] == designation
                    ]
                    td_lignes.append({
                        **periode,
                        "E_MODE": mode,
                        "D_DESIGNATION": designation,
                        "D_MONTANT": sum((decimal_cellule(ligne.get("D_MONTANT")) for ligne in selection), start=Decimal("0")),
                    })
        td_entetes = ["AJ_ANNEE_Z", "AJ_MOIS_Z", "E_MODE", "D_DESIGNATION", "D_MONTANT"]
        td = ajouter_feuille(classeur, noms[3], td_entetes, td_lignes, "Z1TotalMontant")
        self.noter_filiation(noms[3], [noms[1]], "agrégation par mois, mode et désignation")
        entetes_modes: dict[str, list[str]] = {}
        feuilles_modes: dict[str, Any] = {}
        for index, mode in enumerate(("ZZ1", "ZZ2", "Z")):
            feuilles_modes[mode], entetes_modes[mode] = self.ajouter_mode_large(
                classeur, noms[4 + index], mois, mode, CIBLES_Z1, td, False, f"Z1Mode{mode}",
            )
            self.noter_filiation(noms[4 + index], [noms[3]], f"copie en valeur filtrée sur le mode {mode}")
        entetes_compare = ["AJ_ANNEE_Z", "AJ_MOIS_Z", "CA_TTC_Z1", "CA_TTC_EJ", "AJ_ECART_CA_TTC", "HORS_TAXE_1_Z1", "HORS_TAXE_1_EJ", "AJ_ECART_HORS_TAXE_1", "TVA1_Z1", "TVA1_EJ", "AJ_ECART_TVA1"]
        definition_ej = resoudre_classeur_751("ej_entetes", boutique=boutique)
        feuille_ej = definition_ej.noms_feuilles(ALIAS_COURT)[12]
        ej = self.lire_feuille_produite(definition_ej.nom_fichier, feuille_ej)
        modes = {ligne["AJ_MOIS_Z"]: ligne for ligne in lire_lignes_feuille(feuilles_modes[mode_retenu])}
        lignes_compare = []
        for periode in mois:
            etiquette = periode["AJ_MOIS_Z"]
            mode_ligne = modes[etiquette]
            valeurs = {
                "CA_TTC_Z1": decimal_cellule(mode_ligne["CA_BRUT_D_MONTANT"]),
                "CA_TTC_EJ": sommer_periode(ej, etiquette, "SOMME_E_TTC"),
                "HORS_TAXE_1_Z1": decimal_cellule(mode_ligne["HORS_TAXE_1_D_MONTANT"]),
                "HORS_TAXE_1_EJ": sommer_periode(ej, etiquette, "SOMME_AJ_TOTAL_HT"),
                "TVA1_Z1": decimal_cellule(mode_ligne["TVA_1_D_MONTANT"]),
                "TVA1_EJ": sommer_periode(ej, etiquette, "SOMME_AJ_TOTAL_TVA_20"),
            }
            lignes_compare.append({
                **periode, **valeurs,
                "AJ_ECART_CA_TTC": valeurs["CA_TTC_Z1"] - valeurs["CA_TTC_EJ"],
                "AJ_ECART_HORS_TAXE_1": valeurs["HORS_TAXE_1_Z1"] - valeurs["HORS_TAXE_1_EJ"],
                "AJ_ECART_TVA1": valeurs["TVA1_Z1"] - valeurs["TVA1_EJ"],
            })
        compare = ajouter_feuille(classeur, noms[7], entetes_compare, lignes_compare, "CompareZ1EJ")
        self.noter_filiation(noms[7], [feuilles_modes[mode_retenu].title, f"{definition_ej.nom_fichier}#{feuille_ej}"], "comparaison Z1/EJ")
        for ecart in (4, 7, 10):
            mise_en_forme_ecart(compare, ecart, len(mois))
        self.exporter(classeur, definition)

    def construire_recettes(self) -> None:
        classeur = nouveau_classeur()
        self.commencer_filiations()
        definition = resoudre_classeur_751("recettes_toutes")
        sources: dict[str, tuple[str, str, dict[str, dict[str, Any]]]] = {}
        for boutique in BOUTIQUES:
            definition_ej = resoudre_classeur_751("ej_entetes", boutique=boutique)
            feuille_source = definition_ej.noms_feuilles(ALIAS_COURT)[12]
            lignes_source = self.lire_feuille_produite(definition_ej.nom_fichier, feuille_source)
            sources[boutique] = (
                definition_ej.nom_fichier,
                feuille_source,
                {str(ligne["AJ_MOIS"]): ligne for ligne in lignes_source},
            )
        lignes = [{
            **mois,
            "MASSENA_SOMME_AJ_TOTAL_HT": decimal_cellule(sources["MASSENA"][2][mois["AJ_MOIS"]]["SOMME_AJ_TOTAL_HT"]),
            "MASSENA_SOMME_AJ_TOTAL_TVA_20": decimal_cellule(sources["MASSENA"][2][mois["AJ_MOIS"]]["SOMME_AJ_TOTAL_TVA_20"]),
            "MASSENA_SOMME_E_TTC": decimal_cellule(sources["MASSENA"][2][mois["AJ_MOIS"]]["SOMME_E_TTC"]),
            "MATURIN_SOMME_AJ_TOTAL_HT": decimal_cellule(sources["MATURIN"][2][mois["AJ_MOIS"]]["SOMME_AJ_TOTAL_HT"]),
            "MATURIN_SOMME_AJ_TOTAL_TVA_20": decimal_cellule(sources["MATURIN"][2][mois["AJ_MOIS"]]["SOMME_AJ_TOTAL_TVA_20"]),
            "MATURIN_SOMME_E_TTC": decimal_cellule(sources["MATURIN"][2][mois["AJ_MOIS"]]["SOMME_E_TTC"]),
        } for mois in lignes_mensuelles()]
        for ligne in lignes:
            ligne["AJ_TOTAL_TOUS_BOUTIQUE_HT"] = ligne["MASSENA_SOMME_AJ_TOTAL_HT"] + ligne["MATURIN_SOMME_AJ_TOTAL_HT"]
            ligne["AJ_TOTAL_TOUS_BOUTIQUE_TVA"] = ligne["MASSENA_SOMME_AJ_TOTAL_TVA_20"] + ligne["MATURIN_SOMME_AJ_TOTAL_TVA_20"]
            ligne["AJ_TOTAL_TOUS_BOUTIQUE_TTC"] = ligne["MASSENA_SOMME_E_TTC"] + ligne["MATURIN_SOMME_E_TTC"]
        entetes = ["AJ_ANNEE", "AJ_MOIS", "MASSENA_SOMME_AJ_TOTAL_HT", "MASSENA_SOMME_AJ_TOTAL_TVA_20", "MASSENA_SOMME_E_TTC", "MATURIN_SOMME_AJ_TOTAL_HT", "MATURIN_SOMME_AJ_TOTAL_TVA_20", "MATURIN_SOMME_E_TTC", "AJ_TOTAL_TOUS_BOUTIQUE_HT", "AJ_TOTAL_TOUS_BOUTIQUE_TVA", "AJ_TOTAL_TOUS_BOUTIQUE_TTC"]
        nom_feuille = definition.noms_feuilles(ALIAS_COURT)[0]
        ajouter_feuille(classeur, nom_feuille, entetes, lignes, "RecettesTous")
        self.noter_filiation(nom_feuille, [f"{sources[boutique][0]}#{sources[boutique][1]}" for boutique in BOUTIQUES], "assemblage en valeur des deux boutiques")
        self.exporter(classeur, definition)

    def construire_ca3(self) -> None:
        classeur = nouveau_classeur()
        self.commencer_filiations()
        definition = resoudre_classeur_751("comparaison_ca3")
        definition_recettes = resoudre_classeur_751("recettes_toutes")
        feuille_recettes = definition_recettes.noms_feuilles(ALIAS_COURT)[0]
        recettes = self.lire_feuille_produite(definition_recettes.nom_fichier, feuille_recettes)
        lignes = [{
            "AJ_ANNEE": ligne["AJ_ANNEE"],
            "AJ_MOIS": ligne["AJ_MOIS"],
            "AJ_TOTAL_TOUS_BOUTIQUE_HT": decimal_cellule(ligne["AJ_TOTAL_TOUS_BOUTIQUE_HT"]),
            "AJ_TOTAL_TOUS_BOUTIQUE_TVA": decimal_cellule(ligne["AJ_TOTAL_TOUS_BOUTIQUE_TVA"]),
            "AJ_TOTAL_TOUS_BOUTIQUE_TTC": decimal_cellule(ligne["AJ_TOTAL_TOUS_BOUTIQUE_TTC"]),
        } for ligne in recettes]
        entetes = ["AJ_ANNEE", "AJ_MOIS", "AJ_TOTAL_TOUS_BOUTIQUE_HT", "AJ_TOTAL_TOUS_BOUTIQUE_TVA", "AJ_TOTAL_TOUS_BOUTIQUE_TTC", "MTT_HT1_CA3", "MTT_HT1_20_CA3", "MTT_TVA_20_CA3", "AJ_ECART_HT20", "AJ_ECART_TVA20", "STATUT", "COMMENTAIRE"]
        nom_feuille = definition.noms_feuilles(ALIAS_COURT)[0]
        feuille = ajouter_feuille(classeur, nom_feuille, entetes, lignes, "CompareCA3")
        self.noter_filiation(nom_feuille, [f"{definition_recettes.nom_fichier}#{feuille_recettes}", "déclarations CA3 à fournir"], "copie des recettes reconstituées et comparaison CA3")
        colonne_formules(feuille, 8, len(lignes), lambda r, _i: f'=IF(G{r}="","",C{r}-G{r})')
        colonne_formules(feuille, 9, len(lignes), lambda r, _i: f'=IF(H{r}="","",D{r}-H{r})')
        colonne_formules(feuille, 10, len(lignes), lambda r, _i: f'=IF(COUNTA(F{r}:H{r})=0,"À SAISIR","RENSEIGNÉ")')
        for ligne in range(2, len(lignes) + 2):
            for colonne in range(6, 9):
                feuille.cell(ligne, colonne).fill = copy(REMPLISSAGE_SAISIE)
            feuille.cell(ligne, 12).fill = copy(REMPLISSAGE_SAISIE)
        feuille.conditional_formatting.add(f"K2:K{len(lignes)+1}", FormulaRule(formula=['ISNUMBER(SEARCH("À SAISIR",K2))'], fill=REMPLISSAGE_ALERTE, font=POLICE_ALERTE))
        mise_en_forme_ecart(feuille, 8, len(lignes))
        mise_en_forme_ecart(feuille, 9, len(lignes))

        entetes_fec = ["EXERCICE", "PERIODE", "GESCo_HT", "GESCo_TVA", "GESCo_TTC", "CA_HT_FEC", "TVA_COLLECTEE_FEC", "CA_TTC_FEC", "AJ_ECART_HT", "AJ_ECART_TVA", "AJ_ECART_TTC", "SOURCE_FEC", "STATUT"]
        lignes_fec = []
        for annee in ANNEES:
            selection = [ligne for ligne in lignes if ligne["AJ_ANNEE"] == annee]
            lignes_fec.append([
                annee, "2025-01-01 au 2025-08-31" if annee == 2025 else f"{annee}-01-01 au {annee}-12-31",
                sum((ligne["AJ_TOTAL_TOUS_BOUTIQUE_HT"] for ligne in selection), Decimal("0")),
                sum((ligne["AJ_TOTAL_TOUS_BOUTIQUE_TVA"] for ligne in selection), Decimal("0")),
                sum((ligne["AJ_TOTAL_TOUS_BOUTIQUE_TTC"] for ligne in selection), Decimal("0")),
                None, None, None, None, None, None, None, None,
            ])
        for colonne, valeur in enumerate(entetes_fec, start=14):
            cellule = feuille.cell(1, colonne, valeur)
            cellule.fill = copy(REMPLISSAGE_ENTETE)
            cellule.font = Font(name="Aptos Display", size=10, bold=True, color="FFFFFF")
            cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cellule.border = Border(top=BORD_ENTETE, bottom=BORD_ENTETE, left=BORD_ENTETE, right=BORD_ENTETE)
        for index_ligne, valeurs in enumerate(lignes_fec, start=2):
            for index_colonne, valeur in enumerate(valeurs, start=14):
                cellule = feuille.cell(index_ligne, index_colonne, valeur)
                cellule.font = Font(name="Aptos", size=9)
                cellule.border = Border(bottom=BORD_FIN)
                if 16 <= index_colonne <= 24:
                    cellule.number_format = FORMAT_MONTANT
            for colonne in range(19, 22):
                feuille.cell(index_ligne, colonne).fill = copy(REMPLISSAGE_SAISIE)
            feuille.cell(index_ligne, 25).fill = copy(REMPLISSAGE_SAISIE)
            feuille.cell(index_ligne, 22, f'=IF(S{index_ligne}="","",P{index_ligne}-S{index_ligne})')
            feuille.cell(index_ligne, 23, f'=IF(T{index_ligne}="","",Q{index_ligne}-T{index_ligne})')
            feuille.cell(index_ligne, 24, f'=IF(U{index_ligne}="","",R{index_ligne}-U{index_ligne})')
            feuille.cell(index_ligne, 26, f'=IF(COUNTA(S{index_ligne}:U{index_ligne})=0,"À SAISIR","RENSEIGNÉ")')
        feuille.column_dimensions["O"].width = 26
        for lettre in ("P", "Q", "R", "S", "T", "U", "V", "W", "X"):
            feuille.column_dimensions[lettre].width = 17
        feuille.column_dimensions["Y"].width = 24
        feuille.column_dimensions["Z"].width = 16
        feuille.conditional_formatting.add("Z2:Z4", FormulaRule(formula=['ISNUMBER(SEARCH("À SAISIR",Z2))'], fill=REMPLISSAGE_ALERTE, font=POLICE_ALERTE))
        for colonne in (21, 22, 23):
            mise_en_forme_ecart(feuille, colonne, 3)
        self.exporter(classeur, definition)

    def rendre_qa(self) -> int:
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        pdftoppm = shutil.which("pdftoppm")
        if not soffice or not pdftoppm:
            raise FileNotFoundError("Le mode --qa nécessite LibreOffice (soffice) et pdftoppm.")
        if self.qa_dir.exists():
            shutil.rmtree(self.qa_dir)
        self.qa_dir.mkdir(parents=True)
        pages = 0
        for entree in self.manifeste["workbooks"]:
            nom = entree["fileName"]
            destination = self.qa_dir / Path(nom).stem
            destination.mkdir()
            with tempfile.TemporaryDirectory(prefix="qa_751_") as temporaire:
                temporaire_path = Path(temporaire)
                source = self.xlsx / nom
                classeur = load_workbook(source)
                for feuille in classeur.worksheets:
                    derniere_colonne = min(max(feuille.max_column, 1), 26)
                    derniere_ligne = min(max(feuille.max_row, 1), 16)
                    feuille.print_area = f"A1:{get_column_letter(derniere_colonne)}{derniere_ligne}"
                    feuille.page_setup.orientation = "landscape"
                    feuille.page_setup.fitToWidth = 1
                    feuille.page_setup.fitToHeight = 1
                    feuille.sheet_properties.pageSetUpPr.fitToPage = True
                    feuille.sheet_properties.pageSetUpPr.autoPageBreaks = False
                temporaire_xlsx = temporaire_path / nom
                classeur.save(temporaire_xlsx)
                classeur.close()
                profil_libreoffice = temporaire_path / "profil_libreoffice"
                conversion = subprocess.run(
                    [
                        soffice,
                        f"-env:UserInstallation={profil_libreoffice.as_uri()}",
                        "--headless",
                        "--convert-to",
                        "pdf",
                        "--outdir",
                        str(temporaire_path),
                        str(temporaire_xlsx),
                    ],
                    check=False,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                )
                if conversion.returncode:
                    raise RuntimeError(
                        f"Échec du rendu LibreOffice pour {nom} "
                        f"(code {conversion.returncode}) : {conversion.stdout.strip()}"
                    )
                pdf = temporaire_path / f"{Path(nom).stem}.pdf"
                if not pdf.is_file():
                    raise RuntimeError(f"Aperçu PDF non produit pour {nom}")
                prefixe = destination / "feuille"
                subprocess.run([pdftoppm, "-png", "-r", "90", str(pdf), str(prefixe)], check=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
                images = sorted(destination.glob("feuille-*.png"))
                if len(images) != entree["sheetCount"]:
                    raise RuntimeError(f"{nom}: {len(images)} aperçus pour {entree['sheetCount']} feuilles")
                pages += len(images)
        return pages

    def construire(self) -> dict[str, Any]:
        for boutique in BOUTIQUES:
            self.construire_entetes(boutique)
            self.construire_lignes(boutique)
        for boutique in BOUTIQUES:
            for annee in ANNEES:
                self.construire_z1(boutique, annee)
                self.construire_z2(boutique, annee)
        self.construire_recettes()
        self.construire_ca3()
        self.manifeste["summary"] = {
            "excelFiles": len(self.manifeste["workbooks"]),
            "excelSheets": sum(entree["sheetCount"] for entree in self.manifeste["workbooks"]),
            "csvFiles": len(self.manifeste["csvFiles"]),
        }
        self.manifeste["controle"] = json.loads((self.controle / "RESUME_EXPORT_751.json").read_text(encoding="utf-8"))
        if self.qa:
            self.manifeste["summary"]["qaPreviews"] = self.rendre_qa()
        for entree in self.manifeste["workbooks"]:
            chemin = self.xlsx / entree["fileName"]
            entree["sha256"] = hashlib.sha256(chemin.read_bytes()).hexdigest()
        self.controle.mkdir(parents=True, exist_ok=True)
        (self.controle / "MANIFESTE_CONTROLE_751.json").write_text(
            json.dumps(self.manifeste, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return self.manifeste["summary"]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--staging", type=Path, required=True)
    parser.add_argument("--controle", type=Path, required=True)
    parser.add_argument("--sortie", type=Path, required=True)
    parser.add_argument("--qa-dir", type=Path, required=True)
    parser.add_argument("--qa", action="store_true")
    args = parser.parse_args(argv)
    resume = Constructeur751(
        args.staging.resolve(),
        args.controle.resolve(),
        args.sortie.resolve(),
        args.qa_dir.resolve(),
        args.qa,
    ).construire()
    print(json.dumps(resume, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
